"""
Excel Parsers for DVT Test Report Generator
Dictionary-based parsers for AI-friendly data processing
"""

import openpyxl
from typing import Dict, List, Any, Optional, Set
import tempfile
import os
import re


class BaseExcelParser:
    """
    Base class for Excel parsers with dictionary-based data structure for AI compatibility
    """
    
    def __init__(self):
        self.parsed_data = {}
    
    async def parse_uploaded_file(self, uploaded_file) -> Dict[str, Dict[str, Any]]:
        """
        Parse uploaded file object
        
        Args:
            uploaded_file: FastAPI UploadFile or similar file object
            
        Returns:
            Dictionary with parsed sheet data
        """
        try:
            # Reset file pointer and read content
            if hasattr(uploaded_file, 'seek') and hasattr(uploaded_file, 'read'):
                await uploaded_file.seek(0)
                content = await uploaded_file.read()
            elif hasattr(uploaded_file, 'file'):
                uploaded_file.file.seek(0)
                content = uploaded_file.file.read()
            else:
                raise ValueError("Unsupported file object type")
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_file.write(content)
                temp_file.flush()
                
                # Parse the temporary file
                results = self.parse_excel_file(temp_file.name)
                
                # Cleanup
                os.unlink(temp_file.name)
                
                return results
                
        except Exception as e:
            print(f"❌ Error parsing uploaded file: {str(e)}")
            return {}
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Parse Excel file - to be implemented by subclasses
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with sheet data
        """
        raise NotImplementedError("Subclasses must implement parse_excel_file")
    
    def _parse_sheet_to_dict(self, worksheet, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Parse worksheet to dictionary format for AI processing
        
        Args:
            worksheet: openpyxl worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary with sheet data using dict format for each row
        """
        try:
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row < 1:
                print(f"⚠️ Sheet {sheet_name} is empty")
                return None
            
            # Get column names from first row
            columns = []
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                column_name = str(cell_value).strip() if cell_value else f"Column{col}"
                columns.append(column_name)
            
            # Get data rows as dictionaries (skip header row)
            data = []
            data_row_count = 0
            
            for row in range(2, max_row + 1):  # Start from row 2 (skip header)
                row_dict = {}
                has_data = False
                
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    cell_str = str(cell_value).strip() if cell_value else ""
                    
                    # Use column name as key
                    column_name = columns[col-1] if (col-1) < len(columns) else f"Column{col}"
                    row_dict[column_name] = cell_str
                    
                    if cell_str:  # Check if row has any data
                        has_data = True
                
                # Only include rows that have some data
                if has_data:
                    data.append(row_dict)
                    data_row_count += 1
            
            return {
                "columns": columns,
                "row_count": data_row_count,
                "data": data,  # Now each item is a dictionary
                "sheet_name": sheet_name
            }
            
        except Exception as e:
            print(f"❌ Error parsing sheet {sheet_name}: {str(e)}")
            return None


class DeviationsParser(BaseExcelParser):
    """
    Parser for DEVIATIONS sheet
    Uses dictionary format for better AI comprehension
    """
    
    def __init__(self):
        super().__init__()
        self.target_sheets = ["DEVIATIONS", "PROTOCOL DEVIATIONS", "DEVIATION LOG"]
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Parse Excel file and extract data from deviations sheet
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with deviations data using AI-friendly dictionary format
        """
        results = {}
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print(f"📋 DeviationsParser: Found sheets: {workbook.sheetnames}")
            
            for sheet_name in workbook.sheetnames:
                # Check if this is one of our target sheets
                normalized_name = self._normalize_sheet_name(sheet_name)
                if normalized_name in self.target_sheets:
                    sheet_data = self._parse_sheet_to_dict(workbook[sheet_name], normalized_name)
                    if sheet_data:
                        results[normalized_name] = sheet_data
                        print(f"✅ Parsed {normalized_name}: {sheet_data['row_count']} deviations, {len(sheet_data['columns'])} columns")
                        if sheet_data['data']:
                            print(f"   Sample deviation fields: {list(sheet_data['data'][0].keys())[:3]}...")
        
        except Exception as e:
            print(f"❌ Error parsing Excel file for deviations: {str(e)}")
        
        self.parsed_data = results
        return results
    
    def _normalize_sheet_name(self, sheet_name: str) -> str:
        """Normalize sheet name to match target sheets"""
        sheet_upper = sheet_name.upper().strip()
        
        # Direct matches
        if sheet_upper in self.target_sheets:
            return sheet_upper
        
        # Partial matches
        if "DEVIATION" in sheet_upper:
            if "PROTOCOL" in sheet_upper:
                return "PROTOCOL DEVIATIONS"
            elif "LOG" in sheet_upper:
                return "DEVIATION LOG"
            else:
                return "DEVIATIONS"
        
        return sheet_name  # Return original if no match
    
    def get_summary(self) -> Dict[str, Dict[str, Any]]:
        """Get summary information for all parsed deviation sheets"""
        summaries = {}
        for sheet_name, data in self.parsed_data.items():
            summaries[sheet_name] = {
                "sheet_name": sheet_name,
                "columns": data["columns"],
                "deviation_count": data["row_count"],
                "column_count": len(data["columns"]),
                "sample_deviation": data["data"][0] if data["data"] else {}
            }
        return summaries
    
    def get_deviations_by_type(self, deviation_type: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Get deviations filtered by type
        
        Args:
            deviation_type: Optional filter by deviation type/severity
            
        Returns:
            List of deviation records
        """
        all_deviations = []
        for sheet_data in self.parsed_data.values():
            for deviation in sheet_data['data']:
                if deviation_type is None:
                    all_deviations.append(deviation)
                else:
                    # Check if deviation matches type (look in common type columns)
                    type_columns = ['Type', 'Deviation Type', 'Category', 'Severity']
                    for col in type_columns:
                        if col in deviation and deviation_type.lower() in deviation[col].lower():
                            all_deviations.append(deviation)
                            break
        return all_deviations
    
    def format_for_ai_prompt(self) -> str:
        """
        Format parsed deviations data as text for AI prompts
        
        Returns:
            Formatted string for AI consumption
        """
        if not self.parsed_data:
            return "No deviations data found."
        
        formatted_sections = []
        
        for sheet_name, sheet_data in self.parsed_data.items():
            formatted_sections.append(f"\n=== {sheet_name} ===")
            formatted_sections.append(f"Found {sheet_data['row_count']} deviations")
            formatted_sections.append(f"Columns: {', '.join(sheet_data['columns'])}")
            
            # Show first few deviation records as examples
            for i, deviation in enumerate(sheet_data['data'][:3], 1):
                formatted_sections.append(f"  Deviation {i}: {deviation}")
            
            if sheet_data['row_count'] > 3:
                formatted_sections.append(f"  ... and {sheet_data['row_count'] - 3} more deviations")
        
        return '\n'.join(formatted_sections)


class DefectiveUnitsParser(BaseExcelParser):
    """
    Parser for DEFECTIVE UNITS sheet
    Uses dictionary format for better AI comprehension
    """
    
    def __init__(self):
        super().__init__()
        self.target_sheets = ["DEFECTIVE UNITS", "DEFECTIVE UNIT INVESTIGATIONS", "FAILED UNITS"]
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Parse Excel file and extract data from defective units sheet
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with defective units data using AI-friendly dictionary format
        """
        results = {}
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print(f"📋 DefectiveUnitsParser: Found sheets: {workbook.sheetnames}")
            
            for sheet_name in workbook.sheetnames:
                # Check if this is one of our target sheets
                normalized_name = self._normalize_sheet_name(sheet_name)
                if normalized_name in self.target_sheets:
                    sheet_data = self._parse_sheet_to_dict(workbook[sheet_name], normalized_name)
                    if sheet_data:
                        results[normalized_name] = sheet_data
                        print(f"✅ Parsed {normalized_name}: {sheet_data['row_count']} defective units, {len(sheet_data['columns'])} columns")
                        if sheet_data['data']:
                            print(f"   Sample defective unit fields: {list(sheet_data['data'][0].keys())[:3]}...")
        
        except Exception as e:
            print(f"❌ Error parsing Excel file for defective units: {str(e)}")
        
        self.parsed_data = results
        return results
    
    def _normalize_sheet_name(self, sheet_name: str) -> str:
        """Normalize sheet name to match target sheets"""
        sheet_upper = sheet_name.upper().strip()
        
        # Direct matches
        if sheet_upper in self.target_sheets:
            return sheet_upper
        
        # Partial matches
        if "DEFECTIVE" in sheet_upper and "UNIT" in sheet_upper:
            if "INVESTIGATION" in sheet_upper:
                return "DEFECTIVE UNIT INVESTIGATIONS"
            else:
                return "DEFECTIVE UNITS"
        elif "FAILED" in sheet_upper and "UNIT" in sheet_upper:
            return "FAILED UNITS"
        
        return sheet_name  # Return original if no match
    
    def get_summary(self) -> Dict[str, Dict[str, Any]]:
        """Get summary information for all parsed defective units sheets"""
        summaries = {}
        for sheet_name, data in self.parsed_data.items():
            summaries[sheet_name] = {
                "sheet_name": sheet_name,
                "columns": data["columns"],
                "defective_unit_count": data["row_count"],
                "column_count": len(data["columns"]),
                "sample_unit": data["data"][0] if data["data"] else {}
            }
        return summaries
    
    def get_units_by_failure_type(self, failure_type: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Get defective units filtered by failure type
        
        Args:
            failure_type: Optional filter by failure type/mode
            
        Returns:
            List of defective unit records
        """
        all_units = []
        for sheet_data in self.parsed_data.values():
            for unit in sheet_data['data']:
                if failure_type is None:
                    all_units.append(unit)
                else:
                    # Check if unit matches failure type (look in common failure columns)
                    failure_columns = ['Failure Type', 'Failure Mode', 'Root Cause', 'Issue Type']
                    for col in failure_columns:
                        if col in unit and failure_type.lower() in unit[col].lower():
                            all_units.append(unit)
                            break
        return all_units
    
    def get_units_by_serial_number(self, serial_numbers: List[str]) -> List[Dict[str, Any]]:
        """
        Get defective units by serial number list
        
        Args:
            serial_numbers: List of serial numbers to find
            
        Returns:
            List of matching defective unit records
        """
        matching_units = []
        for sheet_data in self.parsed_data.values():
            for unit in sheet_data['data']:
                # Check common serial number column names
                serial_columns = ['Serial Number', 'DUT Serial Number', 'Unit Serial', 'SN']
                for col in serial_columns:
                    if col in unit and unit[col] in serial_numbers:
                        matching_units.append(unit)
                        break
        return matching_units
    
    def format_for_ai_prompt(self) -> str:
        """
        Format parsed defective units data as text for AI prompts
        
        Returns:
            Formatted string for AI consumption
        """
        if not self.parsed_data:
            return "No defective units data found."
        
        formatted_sections = []
        
        for sheet_name, sheet_data in self.parsed_data.items():
            formatted_sections.append(f"\n=== {sheet_name} ===")
            formatted_sections.append(f"Found {sheet_data['row_count']} defective units")
            formatted_sections.append(f"Columns: {', '.join(sheet_data['columns'])}")
            
            # Show first few defective unit records as examples
            for i, unit in enumerate(sheet_data['data'][:3], 1):
                formatted_sections.append(f"  Defective Unit {i}: {unit}")
            
            if sheet_data['row_count'] > 3:
                formatted_sections.append(f"  ... and {sheet_data['row_count'] - 3} more defective units")
        
        return '\n'.join(formatted_sections)


class EquipmentUsedParser(BaseExcelParser):
    """
    Parser for equipment used Excel sheets: EQUIPMENT LOG, SOFTWARE LOG, MATERIAL LOG
    Uses dictionary format for better AI comprehension
    """
    
    def __init__(self):
        super().__init__()
        self.target_sheets = ["EQUIPMENT LOG", "SOFTWARE LOG", "MATERIAL LOG"]
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Parse Excel file and extract data from target sheets
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with sheet data using AI-friendly dictionary format
        """
        results = {}
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print(f"📋 EquipmentUsedParser: Found sheets: {workbook.sheetnames}")
            
            for sheet_name in workbook.sheetnames:
                # Check if this is one of our target sheets
                normalized_name = self._normalize_sheet_name(sheet_name)
                if normalized_name in self.target_sheets:
                    sheet_data = self._parse_sheet_to_dict(workbook[sheet_name], normalized_name)
                    if sheet_data:
                        results[normalized_name] = sheet_data
                        print(f"✅ Parsed {normalized_name}: {sheet_data['row_count']} rows, {len(sheet_data['columns'])} columns")
                        if sheet_data['data']:
                            print(f"   Sample data: {list(sheet_data['data'][0].keys())[:3]}...")
        
        except Exception as e:
            print(f"❌ Error parsing Excel file: {str(e)}")
        
        self.parsed_data = results
        return results
    
    def _normalize_sheet_name(self, sheet_name: str) -> str:
        """Normalize sheet name to match target sheets"""
        sheet_upper = sheet_name.upper().strip()
        
        # Direct matches
        if sheet_upper in self.target_sheets:
            return sheet_upper
        
        # Partial matches
        if "EQUIPMENT" in sheet_upper and "LOG" in sheet_upper:
            return "EQUIPMENT LOG"
        elif "SOFTWARE" in sheet_upper and "LOG" in sheet_upper:
            return "SOFTWARE LOG"
        elif "MATERIAL" in sheet_upper and "LOG" in sheet_upper:
            return "MATERIAL LOG"
        
        return sheet_name  # Return original if no match
    
    def get_summary(self) -> Dict[str, Dict[str, Any]]:
        """Get summary information for all parsed sheets"""
        summaries = {}
        for sheet_name, data in self.parsed_data.items():
            summaries[sheet_name] = {
                "sheet_name": sheet_name,
                "columns": data["columns"],
                "row_count": data["row_count"],
                "column_count": len(data["columns"]),
                "sample_record": data["data"][0] if data["data"] else {}
            }
        return summaries
    
    def format_for_ai_prompt(self) -> str:
        """
        Format parsed data as text for AI prompts
        
        Returns:
            Formatted string for AI consumption
        """
        if not self.parsed_data:
            return "No equipment, software, or material log data found."
        
        formatted_sections = []
        
        for sheet_name, sheet_data in self.parsed_data.items():
            formatted_sections.append(f"\n=== {sheet_name} ===")
            formatted_sections.append(f"Found {sheet_data['row_count']} items")
            formatted_sections.append(f"Columns: {', '.join(sheet_data['columns'])}")
            
            # Show first few records as examples
            for i, record in enumerate(sheet_data['data'][:3], 1):
                formatted_sections.append(f"  Item {i}: {record}")
            
            if sheet_data['row_count'] > 3:
                formatted_sections.append(f"  ... and {sheet_data['row_count'] - 3} more items")
        
        return '\n'.join(formatted_sections)


class TestArticleParser(BaseExcelParser):
    """
    Parser for TEST ARTICLE LOG & TEST RESULTS sheet
    Uses DUT Serial Number as primary key with dictionary format for AI processing
    """
    
    def __init__(self):
        super().__init__()
        self.target_sheet_patterns = [
            "TEST ARTICLE LOG & TEST RESULTS",
            "TEST ARTICLE LOG",
            "TEST RESULTS"
        ]
        self.primary_key = "DUT Serial Number"
        
    def parse_excel_file(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Parse Excel file for TEST ARTICLE LOG & TEST RESULTS data
        
        Returns:
            Dictionary with parsed data
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print(f"📋 TestArticleParser: Found sheets: {workbook.sheetnames}")
            
            # Find target sheet
            target_sheet = None
            sheet_name = None
            
            for sheet in workbook.sheetnames:
                if self._is_target_sheet(sheet):
                    target_sheet = workbook[sheet]
                    sheet_name = sheet
                    print(f"✅ Found target sheet: {sheet_name}")
                    break
            
            if not target_sheet or not sheet_name:
                print("⚠️ No TEST ARTICLE LOG & TEST RESULTS sheet found")
                return {}
            
            # Parse sheet data
            sheet_data = self._parse_test_article_sheet(target_sheet, sheet_name)
            
            if sheet_data:
                result = {"TEST_ARTICLE_DATA": sheet_data}
                self.parsed_data = result
                return result
            else:
                return {}
                
        except Exception as e:
            print(f"❌ Error parsing test article data: {str(e)}")
            return {}
    
    def _is_target_sheet(self, sheet_name: str) -> bool:
        """Check if sheet name matches target patterns"""
        sheet_upper = sheet_name.upper()
        return any(pattern.upper() in sheet_upper for pattern in self.target_sheet_patterns)
    
    def _parse_test_article_sheet(self, worksheet, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Parse TEST ARTICLE LOG & TEST RESULTS sheet with DUT Serial Number as key
        
        Returns:
            Dictionary with test article data using AI-friendly format
        """
        try:
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row < 2:
                print(f"⚠️ Sheet {sheet_name} has insufficient data")
                return None
            
            # Get column names from first row
            columns = []
            primary_key_col = None
            
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                column_name = str(cell_value).strip() if cell_value else f"Column{col}"
                columns.append(column_name)
                
                # Find primary key column
                if self.primary_key.lower() in column_name.lower() or "dut serial" in column_name.lower():
                    primary_key_col = col
                    self.primary_key = column_name  # Use actual column name
            
            if primary_key_col is None:
                print(f"⚠️ Primary key '{self.primary_key}' not found in sheet")
                return None
            
            print(f"🔑 Using primary key: {self.primary_key} (column {primary_key_col})")
            print(f"📝 Columns found: {columns}")
            
            # Parse data rows with primary key as dictionary key
            records = {}
            test_result_columns = []
            
            for row in range(2, max_row + 1):
                row_dict = {}
                primary_key_value = None
                has_data = False
                
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    cell_str = str(cell_value).strip() if cell_value else ""
                    
                    column_name = columns[col-1] if (col-1) < len(columns) else f"Column{col}"
                    row_dict[column_name] = cell_str
                    
                    # Track primary key value
                    if col == primary_key_col:
                        primary_key_value = cell_str
                    
                    # Identify test result columns
                    if re.match(r'^test result', column_name.lower()) and column_name not in test_result_columns:
                        test_result_columns.append(column_name)
                    
                    if cell_str:
                        has_data = True
                
                # Store record with primary key as dictionary key
                if has_data and primary_key_value:
                    records[primary_key_value] = row_dict
            
            # Analyze test results
            test_results_analysis = self._analyze_test_results(records, test_result_columns)
            
            result = {
                "sheet_name": sheet_name,
                "primary_key": self.primary_key,
                "columns": columns,
                "total_units": len(records),
                "test_result_columns": test_result_columns,
                "test_results_analysis": test_results_analysis,
                "records": records  # Dictionary with DUT Serial Number as key
            }
            
            print(f"✅ Parsed {len(records)} test article records")
            print(f"📊 Test result columns: {test_result_columns}")
            print(f"📈 Analysis: {test_results_analysis}")
            
            return result
            
        except Exception as e:
            print(f"❌ Error parsing test article sheet: {str(e)}")
            return None
    
    def _analyze_test_results(self, records: Dict[str, Dict], test_result_columns: List[str]) -> Dict[str, Any]:
        """
        Analyze test results for statistical information
        """
        if not test_result_columns:
            return {"message": "No test result columns found"}
        
        analysis = {
            "test_method_losses": 0,
            "total_tests_performed": 0,
            "pass_count": 0,
            "fail_count": 0,
            "tml_count": 0,
            "actual_sample_size": 0
        }
        
        for serial_number, record in records.items():
            unit_has_test_data = False
            
            for col in test_result_columns:
                result = record.get(col, "").upper().strip()
                if result:
                    analysis["total_tests_performed"] += 1
                    unit_has_test_data = True
                    
                    if result == "PASS":
                        analysis["pass_count"] += 1
                    elif result == "FAIL":
                        analysis["fail_count"] += 1
                    elif result in ["TML", "TEST METHOD LOSS"]:
                        analysis["tml_count"] += 1
                        analysis["test_method_losses"] += 1
            
            if unit_has_test_data:
                analysis["actual_sample_size"] += 1
        
        # Calculate actual sample size as total units minus test method losses
        analysis["actual_sample_size"] = len(records) - analysis["test_method_losses"]
        
        return analysis
    
    def get_dut_count(self) -> int:
        """Get total number of DUT units"""
        if "TEST_ARTICLE_DATA" in self.parsed_data:
            return self.parsed_data["TEST_ARTICLE_DATA"]["total_units"]
        return 0
    
    def get_test_results_summary(self) -> Dict[str, Any]:
        """Get summary of test results analysis"""
        if "TEST_ARTICLE_DATA" in self.parsed_data:
            return self.parsed_data["TEST_ARTICLE_DATA"]["test_results_analysis"]
        return {}
    
    def get_records_by_result(self, result_type: str) -> List[Dict[str, Any]]:
        """
        Get all records that have a specific test result
        
        Args:
            result_type: "PASS", "FAIL", "TML", etc.
            
        Returns:
            List of records matching the result type
        """
        if "TEST_ARTICLE_DATA" not in self.parsed_data:
            return []
        
        data = self.parsed_data["TEST_ARTICLE_DATA"]
        matching_records = []
        
        for record in data["records"].values():
            for col in data["test_result_columns"]:
                if record.get(col, "").upper().strip() == result_type.upper():
                    matching_records.append(record)
                    break
        
        return matching_records
    
    def format_for_ai_prompt(self) -> str:
        """
        Format parsed data as text for AI prompts
        
        Returns:
            Formatted string for AI consumption
        """
        if "TEST_ARTICLE_DATA" not in self.parsed_data:
            return "No test article data found."
        
        data = self.parsed_data["TEST_ARTICLE_DATA"]
        
        formatted_lines = [
            f"=== {data['sheet_name']} ===",
            f"Primary Key: {data['primary_key']}",
            f"Total DUT Units: {data['total_units']}",
            f"Columns: {', '.join(data['columns'])}",
            f"Test Result Columns: {', '.join(data['test_result_columns'])}",
            ""
        ]
        
        # Add analysis summary
        analysis = data['test_results_analysis']
        formatted_lines.extend([
            "Test Results Analysis:",
            f"  - Total Tests Performed: {analysis['total_tests_performed']}",
            f"  - PASS: {analysis['pass_count']}",
            f"  - FAIL: {analysis['fail_count']}",
            f"  - Test Method Losses: {analysis['test_method_losses']}",
            f"  - Actual Sample Size: {analysis['actual_sample_size']}",
            ""
        ])
        
        # Add sample records
        formatted_lines.append("Sample Records:")
        for i, (serial_num, record) in enumerate(list(data['records'].items())[:3], 1):
            formatted_lines.append(f"  DUT {i} ({serial_num}): {record}")
        
        if data['total_units'] > 3:
            formatted_lines.append(f"  ... and {data['total_units'] - 3} more DUT units")
        
        return '\n'.join(formatted_lines)


class CombinedExcelParser:
    """
    Combined parser that processes all Excel sheets for the DVT report generator
    """
    
    def __init__(self):
        self.test_article_parser = TestArticleParser()
        self.equipment_used_parser = EquipmentUsedParser()
        self.deviations_parser = DeviationsParser()
        self.defective_units_parser = DefectiveUnitsParser()
        self.all_parsed_data = {}
    
    async def parse_all_excel_files(self, data_files: List[Any]) -> Dict[str, Any]:
        """
        Parse all uploaded Excel files using all parsers
        
        Args:
            data_files: List of uploaded Excel files
            
        Returns:
            Combined dictionary with all parsed data
        """
        print("\n" + "="*60)
        print("📊 EXCEL PARSING STARTED (WITH ALL PARSERS)")
        print("="*60)
        
        all_results = {
            "test_article_data": {},
            "equipment_log_data": {},
            "deviations_data": {},
            "defective_units_data": {},
            "parsing_summary": {
                "files_processed": 0,
                "sheets_found": [],
                "errors": []
            }
        }
        
        for i, data_file in enumerate(data_files):
            if not data_file:
                continue
            
            filename = getattr(data_file, 'filename', f'file_{i}')
            print(f"\n📁 Processing file: {filename}")
            
            try:
                # Parse with TestArticleParser
                print("🔍 Parsing test article data...")
                test_results = await self.test_article_parser.parse_uploaded_file(data_file)
                if test_results:
                    all_results["test_article_data"].update(test_results)
                    all_results["parsing_summary"]["sheets_found"].extend(test_results.keys())
                
                # Parse with EquipmentUsedParser
                print("🔍 Parsing equipment/software/material logs...")
                equipment_results = await self.equipment_used_parser.parse_uploaded_file(data_file)
                if equipment_results:
                    all_results["equipment_log_data"].update(equipment_results)
                    all_results["parsing_summary"]["sheets_found"].extend(equipment_results.keys())
                
                # Parse with DeviationsParser
                print("🔍 Parsing deviations data...")
                deviations_results = await self.deviations_parser.parse_uploaded_file(data_file)
                if deviations_results:
                    all_results["deviations_data"].update(deviations_results)
                    all_results["parsing_summary"]["sheets_found"].extend(deviations_results.keys())
                
                # Parse with DefectiveUnitsParser
                print("🔍 Parsing defective units data...")
                defective_results = await self.defective_units_parser.parse_uploaded_file(data_file)
                if defective_results:
                    all_results["defective_units_data"].update(defective_results)
                    all_results["parsing_summary"]["sheets_found"].extend(defective_results.keys())
                
                all_results["parsing_summary"]["files_processed"] += 1
                
            except Exception as e:
                error_msg = f"Error processing {filename}: {str(e)}"
                print(f"❌ {error_msg}")
                all_results["parsing_summary"]["errors"].append(error_msg)
        
        self.all_parsed_data = all_results
        self._print_parsing_summary()
        
        return all_results
    
    def _print_parsing_summary(self):
        """Print detailed summary of parsing results"""
        print("\n" + "="*60)
        print("📊 EXCEL PARSING SUMMARY (ALL PARSERS)")
        print("="*60)
        
        summary = self.all_parsed_data["parsing_summary"]
        print(f"Files Processed: {summary['files_processed']}")
        print(f"Sheets Found: {', '.join(summary['sheets_found'])}")
        
        if summary['errors']:
            print(f"Errors: {len(summary['errors'])}")
            for error in summary['errors']:
                print(f"  ❌ {error}")
        
        # Print Test Article Summary
        if self.all_parsed_data["test_article_data"]:
            print(f"\n📊 TEST ARTICLE DATA:")
            test_data = list(self.all_parsed_data["test_article_data"].values())[0]
            analysis = test_data["test_results_analysis"]
            print(f"  Total DUT Units: {test_data['total_units']}")
            print(f"  PASS: {analysis['pass_count']}")
            print(f"  FAIL: {analysis['fail_count']}")
            print(f"  Test Method Losses: {analysis['test_method_losses']}")
            print(f"  Actual Sample Size: {analysis['actual_sample_size']}")
        
        # Print Equipment Log Summary
        if self.all_parsed_data["equipment_log_data"]:
            print(f"\n📊 EQUIPMENT/SOFTWARE/MATERIAL LOGS:")
            for sheet_name, data in self.all_parsed_data["equipment_log_data"].items():
                print(f"  {sheet_name}: {data['row_count']} items")
        
        # Print Deviations Summary
        if self.all_parsed_data["deviations_data"]:
            print(f"\n📊 DEVIATIONS DATA:")
            for sheet_name, data in self.all_parsed_data["deviations_data"].items():
                print(f"  {sheet_name}: {data['row_count']} deviations")
        
        # Print Defective Units Summary
        if self.all_parsed_data["defective_units_data"]:
            print(f"\n📊 DEFECTIVE UNITS DATA:")
            for sheet_name, data in self.all_parsed_data["defective_units_data"].items():
                print(f"  {sheet_name}: {data['row_count']} defective units")
        
        print("\n" + "="*60)
    
    def get_test_article_data(self) -> Optional[Dict[str, Any]]:
        """Get parsed test article data"""
        if self.all_parsed_data["test_article_data"]:
            return list(self.all_parsed_data["test_article_data"].values())[0]
        return None
    
    def get_equipment_log_data(self, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Get specific equipment log data"""
        return self.all_parsed_data["equipment_log_data"].get(sheet_name)
    
    def get_all_equipment_logs(self) -> Dict[str, Any]:
        """Get all equipment log data"""
        return self.all_parsed_data["equipment_log_data"]
    
    def get_deviations_data(self) -> Dict[str, Any]:
        """Get all deviations data"""
        return self.all_parsed_data["deviations_data"]
    
    def get_defective_units_data(self) -> Dict[str, Any]:
        """Get all defective units data"""
        return self.all_parsed_data["defective_units_data"]
