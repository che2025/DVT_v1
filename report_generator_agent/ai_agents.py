"""
AI Agents module for DVT Test Report Generator
Implements specialized AI agents for each report generation task (4.1-4.4)
Based on visualization agent pattern with task-specific prompts and schemas
"""

import re
import json
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
from google.genai.types import GenerateContentConfig
from .ai_prompts import DVTPrompts
from .ai_config_settings import AgentConfig, ErrorConfig


@dataclass
class TaskResult:
    """Standard result container for AI tasks"""
    success: bool
    content: str
    metadata: Dict[str, Any]
    error: Optional[str] = None
    attachments: Optional[List[Dict[str, Any]]] = None


class BaseDVTAgent:
    """Base class for all DVT AI agents"""
    
    def __init__(self, client, model_name: Optional[str] = None):
        self.client = client
        self.model_name = model_name or AgentConfig.DEFAULT_MODEL
        self.default_temperature = 0.7
    
    async def generate_content(self, prompt: str, temperature: float = None) -> str:
        """Generate content using AI with error handling"""
        if not self.client:
            raise Exception(ErrorConfig.ERROR_MESSAGES["ai_not_configured"])
        
        try:
            config = GenerateContentConfig(
                temperature=temperature or self.default_temperature,
            )
            
            response = self.client.models.generate_content(
                model=self.model_name,
                contents=[prompt],
                config=config
            )
            
            return response.text if response else ""
        except Exception as e:
            raise Exception(f"{ErrorConfig.ERROR_MESSAGES['generation_failed']}: {str(e)}")


class ScopeAndPurposeAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.1: Create Scope and Purpose
    Specializes in adapting protocol scope/purpose for report format
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("scope_and_purpose")
    
    def _extract_tables_from_protocol(self, protocol_content: str) -> List[Dict[str, Any]]:
        """Extract table data from protocol content"""
        tables = []
        lines = protocol_content.split('\n')
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Look for table starts (lines with multiple | characters)
            if line.count('|') >= 2:
                table_lines = []
                # Collect all consecutive table lines
                while i < len(lines) and lines[i].strip().count('|') >= 2:
                    table_lines.append(lines[i].strip())
                    i += 1
                
                # Parse the table
                if len(table_lines) >= 2:  # At least header and one row
                    try:
                        # Extract headers
                        header_line = table_lines[0]
                        headers = [cell.strip() for cell in header_line.split('|')[1:-1]]
                        
                        # Skip separator line if it exists
                        data_start = 1
                        if len(table_lines) > 1 and all(c in '-|: ' for c in table_lines[1]):
                            data_start = 2
                        
                        # Extract data rows
                        rows = []
                        for table_line in table_lines[data_start:]:
                            if table_line.strip() and '|' in table_line:
                                row_data = [cell.strip() for cell in table_line.split('|')[1:-1]]
                                if len(row_data) == len(headers) and any(cell for cell in row_data):
                                    rows.append(row_data)
                        
                        # Store table data
                        if headers and rows:
                            table_data = {
                                "headers": headers,
                                "rows": rows,
                                "markdown": '\n'.join(table_lines)
                            }
                            tables.append(table_data)
                            
                    except Exception as e:
                        print(f"Error parsing table: {e}")
                        continue
                continue
            i += 1
        
        return tables
    
    async def create_scope_and_purpose(self, protocol_content: str, protocol_number: str, 
                                     project_name: str) -> TaskResult:
        """
        Generate scope and purpose section from protocol content
        
        Args:
            protocol_content: Full protocol document text
            protocol_number: Protocol document number and revision
            project_name: Project name for the report
            
        Returns:
            TaskResult with formatted scope and purpose section
        """
        
        # First, extract any tables from the protocol
        tables = self._extract_tables_from_protocol(protocol_content)
        
        # Filter for scope-related tables (containing Doc ID, REQ ID, Requirement columns)
        scope_tables = []
        for table in tables:
            headers_lower = [h.lower() for h in table["headers"]]
            if any(keyword in ' '.join(headers_lower) for keyword in ['doc id', 'req id', 'requirement', 'doc_id', 'req_id']):
                scope_tables.append(table)
        
        # Generate prompt with table data
        prompt = DVTPrompts.scope_and_purpose_prompt(
            protocol_content, protocol_number, project_name
        )
        
        # If we found scope tables, add them to the prompt
        if scope_tables:
            table_section = "\n\nIMPORTANT - TABLES FOUND IN SCOPE SECTION:\n"
            for i, table in enumerate(scope_tables):
                table_section += f"\nTable {i+1}:\n{table['markdown']}\n"
            table_section += "\nThese tables MUST be included in the SCOPE section exactly as shown above.\n"
            prompt += table_section
        
        try:
            content = await self.generate_content(
                prompt, 
                temperature=AgentConfig.get_temperature("scope_and_purpose")
            )
            
            # If AI didn't include the tables, append them manually
            if scope_tables and not any('|' in line for line in content.split('\n')):
                content += "\n\n"
                for table in scope_tables:
                    content += f"\n{table['markdown']}\n"
            
            # Parse the generated content
            metadata = {
                "protocol_number": protocol_number,
                "project_name": project_name,
                "extraction_method": "ai_adaptation",
                "tables_found": len(scope_tables),
                "tables_included": len(scope_tables) > 0
            }
            
            return TaskResult(
                success=True,
                content=content.strip(),
                metadata=metadata
            )
            
        except Exception as e:
            return TaskResult(
                success=False,
                content="",
                metadata={},
                error=str(e)
            )


class ReferenceAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.2: Create Reference Section
    Specializes in extracting document numbers and creating reference tables
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("reference_section")
    
    async def create_reference_section(self, report_content: str, 
                                     document_metadata: Dict[str, Dict] = None) -> TaskResult:
        """
        Extract document numbers from report and create reference table
        
        Args:
            report_content: Full report content to scan for document numbers
            document_metadata: Optional knowledge base of document metadata
            
        Returns:
            TaskResult with formatted reference table
        """
        
        prompt = DVTPrompts.reference_section_prompt(report_content)
        
        try:
            content = await self.generate_content(
                prompt, 
                temperature=AgentConfig.get_temperature("reference_section")
            )
            
            # Extract document numbers for metadata
            doc_numbers = self._extract_document_numbers(content)
            
            metadata = {
                "documents_found": len(doc_numbers),
                "document_numbers": doc_numbers,
                "extraction_method": "ai_pattern_matching"
            }
            
            return TaskResult(
                success=True,
                content=content.strip(),
                metadata=metadata
            )
            
        except Exception as e:
            return TaskResult(
                success=False,
                content="",
                metadata={},
                error=str(e)
            )
    
    def _extract_document_numbers(self, table_content: str) -> List[str]:
        """Extract document numbers from the generated table"""
        doc_numbers = []
        lines = table_content.split('\n')
        
        for line in lines:
            if '|' in line and not line.strip().startswith('|---'):
                parts = [part.strip() for part in line.split('|')]
                if len(parts) >= 2 and parts[1] and parts[1] != 'Document No.':
                    doc_num = parts[1].strip()
                    if doc_num and doc_num != 'TBD':
                        doc_numbers.append(doc_num)
        
        return doc_numbers


class TestProcedureSummaryAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.4: Create Test Procedure Summary  
    Specializes in creating comprehensive protocol summaries
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("test_procedure")
    
    async def create_test_procedure_summary(self, protocol_content: str, 
                                          protocol_number: str) -> TaskResult:
        """
        Generate test procedure summary from protocol content
        
        Args:
            protocol_content: Full protocol document text
            protocol_number: Protocol document number for reference
            
        Returns:
            TaskResult with formatted test procedure summary (~200 words)
        """
        
        prompt = DVTPrompts.test_procedure_summary_prompt(protocol_content, protocol_number)
        
        try:
            content = await self.generate_content(
                prompt, 
                temperature=AgentConfig.get_temperature("test_procedure")
            )
            
            # Analyze summary for completeness
            word_count = len(content.split())
            elements_covered = self._analyze_summary_elements(content)
            
            metadata = {
                "protocol_number": protocol_number,
                "word_count": word_count,
                "target_word_count": AgentConfig.get_target_spec("procedure_word_count"),
                "elements_covered": elements_covered,
                "completeness_score": len(elements_covered) / AgentConfig.get_target_spec("procedure_elements")
            }
            
            return TaskResult(
                success=True,
                content=content.strip(),
                metadata=metadata
            )
            
        except Exception as e:
            return TaskResult(
                success=False,
                content="",
                metadata={},
                error=str(e)
            )
    
    def _analyze_summary_elements(self, summary: str) -> List[str]:
        """Analyze which required elements are covered in the summary"""
        summary_lower = summary.lower()
        elements_found = []
        
        # Check for conditioning keywords
        if any(word in summary_lower for word in ['condition', 'prep', 'baseline', 'setup', 'initial']):
            elements_found.append('conditioning')
        
        # Check for parameters keywords  
        if any(word in summary_lower for word in ['parameter', 'measure', 'evaluat', 'assess', 'test']):
            elements_found.append('parameters')
        
        # Check for equipment keywords
        if any(word in summary_lower for word in ['equipment', 'instrument', 'device', 'tool', 'system']):
            elements_found.append('equipment')
        
        # Check for monitoring keywords
        if any(word in summary_lower for word in ['monitor', 'track', 'observ', 'record', 'log']):
            elements_found.append('monitoring')
        
        # Check for duration keywords
        if any(word in summary_lower for word in ['duration', 'time', 'period', 'hour', 'day', 'week']):
            elements_found.append('duration')
        
        return elements_found


class AcronymsDefinitionsAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.3: Create Acronyms & Definitions section
    IMPORTANT: This is executed LAST after all other sections are complete
    Specializes in scanning complete reports for acronyms and terms
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("acronyms_definitions")
    
    async def create_acronyms_and_definitions(self, complete_report_content: str,
                                            acronym_knowledge_base: Dict[str, str] = None) -> TaskResult:
        """
        Scan complete report for acronyms and terms, create two sections
        
        Args:
            complete_report_content: Full completed report text to scan
            acronym_knowledge_base: Optional knowledge base of acronym definitions
            
        Returns:
            TaskResult with both 4.1 Acronyms and 4.2 Definitions sections
        """
        
        # Pre-process content to avoid overly long prompts
        processed_content = self._preprocess_content_for_acronyms(complete_report_content)
        
        prompt = DVTPrompts.acronyms_definitions_prompt(processed_content)
        
        try:
            print("📝 Processing acronyms and definitions from report content...")
            content = await self.generate_content(
                prompt, 
                temperature=AgentConfig.get_temperature("acronyms_definitions")
            )
            
            # Extract counts for metadata
            acronyms_found = self._count_table_entries(content, "Acronym")
            definitions_found = self._count_table_entries(content, "Term")
            
            metadata = {
                "acronyms_count": acronyms_found,
                "definitions_count": definitions_found,
                "execution_order": "final",  # This runs last
                "knowledge_base_used": acronym_knowledge_base is not None,
                "content_processed": len(processed_content)
            }
            
            return TaskResult(
                success=True,
                content=content.strip(),
                metadata=metadata
            )
            
        except Exception as e:
            print(f"⚠️ AI processing failed for acronyms section, using fallback template: {str(e)}")
            # Provide a fallback template if AI processing fails
            fallback_content = self._create_fallback_acronyms_section()
            
            return TaskResult(
                success=True,  # Still mark as success with fallback
                content=fallback_content,
                metadata={
                    "acronyms_count": 3,  # Basic template count
                    "definitions_count": 2,  # Basic template count
                    "execution_order": "final",
                    "fallback_used": True,
                    "error_encountered": str(e)
                }
            )
    
    def _count_table_entries(self, content: str, column_name: str) -> int:
        """Count entries in a table by counting rows with the specified column"""
        lines = content.split('\n')
        count = 0
        
        for line in lines:
            if '|' in line and column_name.lower() not in line.lower() and not line.strip().startswith('|---'):
                parts = [part.strip() for part in line.split('|')]
                if len(parts) >= 2 and parts[1] and parts[1] != column_name:
                    count += 1
        
        return count
    
    def _preprocess_content_for_acronyms(self, content: str) -> str:
        """
        Preprocess content to focus on important parts for acronym extraction
        This prevents prompt from being too long and causing timeouts
        """
        # Split content into lines
        lines = content.split('\n')
        
        # Extract key sections that likely contain acronyms and technical terms
        important_lines = []
        
        for line in lines:
            line_clean = line.strip()
            
            # Skip empty lines
            if not line_clean:
                continue
                
            # Include lines with capital letter acronyms (likely to have acronyms)
            if any(word.isupper() and len(word) >= 2 for word in line_clean.split()):
                important_lines.append(line_clean)
                continue
            
            # Include section headers (likely to contain key terms)
            if line_clean.isupper() or line_clean.startswith('#'):
                important_lines.append(line_clean)
                continue
                
            # Include lines with technical terms (containing technical keywords)
            technical_keywords = ['test', 'protocol', 'device', 'measurement', 'configuration', 
                                'equipment', 'procedure', 'specification', 'requirement', 
                                'analysis', 'validation', 'verification', 'compliance']
            if any(keyword in line_clean.lower() for keyword in technical_keywords):
                important_lines.append(line_clean)
                continue
        
        # Join important lines and limit to reasonable length
        processed_content = '\n'.join(important_lines)
        
        # Further truncate if still too long (keep first 4000 chars)
        if len(processed_content) > 4000:
            processed_content = processed_content[:4000] + "\n[Content truncated for processing]"
        
        return processed_content
    
    def _create_fallback_acronyms_section(self) -> str:
        """
        Create a fallback acronyms section when AI processing fails
        """
        return """ACRONYMS & DEFINITIONS

Acronyms
List in this section all acronyms referenced in this document.

| Acronym | Definition |
|---------|------------|
| DVT | Design Verification Testing |
| DUT | Device Under Test |
| TBD | To Be Determined |

Definitions
List all relevant definitions for terms that have a specific meaning for the purpose of this document.

| Term | Definition |
|------|------------|
| Protocol | Test procedure document |
| Conditioning | Environmental preparation phase |
"""


class DeviceUnderTestAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.5: Create Device Under Test Configuration Section
    Analyzes Excel test article data and generates configuration section using AI
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("device_config")
    
    async def create_device_configuration_section(self, data_files: List[Any], 
                                                 device_config: Dict[str, Any], 
                                                 report_config: Dict[str, Any]) -> TaskResult:
        """
        Analyze test article data from Excel and create Device Under Test Configuration section
        
        Args:
            data_files: List of uploaded Excel files
            device_config: User inputs (sterilized, modified, etc.)
            report_config: Report configuration (number, revision, etc.)
            
        Returns:
            TaskResult with AI-generated device configuration section
        """
        try:
            # Extract and prepare Excel data for AI analysis
            excel_data = await self._extract_and_format_excel_data(data_files)
            
            # DEBUG: Print Excel data to check if it's being extracted
            print(f"🔍 [TASK 4.5 DEBUG] Excel data length: {len(excel_data)}")
            print(f"🔍 [TASK 4.5 DEBUG] Number of data files: {len(data_files) if data_files else 0}")
            print("🔍 [TASK 4.5 DEBUG] Excel data content:")
            print("=" * 80)
            print(excel_data)
            print("=" * 80)
            
            # Count DUT first before creating prompt
            test_article_count = self._count_test_articles_in_content(excel_data)
            print(f"🔍 [TASK 4.5 DEBUG] Calculated DUT count: {test_article_count}")
            
            # Create AI prompt with extracted data, user configuration, and calculated count
            prompt = DVTPrompts.device_under_test_prompt(excel_data, device_config, report_config, test_article_count)
            
            # DEBUG: Print prompt info
            print(f"🔍 [TASK 4.5 DEBUG] AI prompt length: {len(prompt)}")
            print(f"🔍 [TASK 4.5 DEBUG] Prompt contains WORKSHEET: {'WORKSHEET' in prompt}")
            print(f"🔍 [TASK 4.5 DEBUG] Device config: {device_config}")
            
            # Generate content using AI
            content = await self.generate_content(
                prompt, 
                temperature=AgentConfig.get_temperature("device_config")
            )
            
            # Parse and validate the generated content (use pre-calculated count)
            # test_article_count already calculated above
            
            # Create attachment if more than 10 test articles
            attachment_info = None
            if test_article_count > 10 and data_files:
                attachment_info = await self._create_test_article_attachment(
                    data_files, report_config
                )
            
            return TaskResult(
                success=True,
                content=content,
                metadata={
                    "test_article_count": test_article_count,
                    "presentation_type": "detailed_table" if test_article_count <= 10 else "summary",
                    "sterilized": device_config.get("units_sterilized", False),
                    "modified": device_config.get("units_modified", False),
                    "attachment_info": attachment_info,
                    "ai_generated": True
                }
            )
            
        except Exception as e:
            # Return fallback content if AI processing fails
            return TaskResult(
                success=False,
                content=self._create_fallback_section(device_config),
                metadata={"error": str(e)},
                error=str(e)
            )
    
    async def _extract_and_format_excel_data(self, data_files: List[Any]) -> str:
        """
        Extract test article data from Excel files and format for AI analysis
        Specifically looks for 'TEST ARTICLE LOG & TEST RESULTS' worksheet
        """
        import openpyxl
        import tempfile
        import os
        
        print(f"🔍 [EXCEL DEBUG] Starting Excel extraction with {len(data_files) if data_files else 0} files")
        formatted_data = ""
        
        if not data_files:
            print("🔍 [EXCEL DEBUG] No data files provided!")
            return "No test article data found in uploaded files."
        
        for i, data_file in enumerate(data_files):
            print(f"🔍 [EXCEL DEBUG] Processing file {i+1}: {getattr(data_file, 'filename', 'Unknown')}")
            
            if not data_file:
                print(f"🔍 [EXCEL DEBUG] File {i+1} is None")
                continue
                
            try:
                # Check if data_file has the correct structure
                print(f"🔍 DEBUG - File object type: {type(data_file)}")
                
                # Reset file pointer and read content
                if hasattr(data_file, 'seek') and hasattr(data_file, 'read'):
                    # This is an UploadFile object
                    await data_file.seek(0)  # Reset file pointer
                    content = await data_file.read()
                elif hasattr(data_file, 'file'):
                    # This has a file attribute
                    data_file.file.seek(0)
                    content = data_file.file.read()
                else:
                    print(f"🔍 DEBUG - Unknown file object structure: {dir(data_file)}")
                    continue
                    
                print(f"🔍 DEBUG - File content size: {len(content)} bytes")
                
                if len(content) == 0:
                    print(f"🔍 DEBUG - Warning: File {getattr(data_file, 'filename', 'unknown')} is empty!")
                    continue
                
                # Save uploaded file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    temp_file.write(content)
                    temp_file_path = temp_file.name
                    print(f"🔍 DEBUG - Saved temp file: {temp_file_path}, size: {len(content)} bytes")
                
                # Load workbook and find the correct worksheet
                workbook = openpyxl.load_workbook(temp_file_path, data_only=True)
                print(f"🔍 DEBUG - Workbook sheets: {workbook.sheetnames}")
                
                # Look for 'TEST ARTICLE LOG & TEST RESULTS' worksheet
                target_sheet = None
                for sheet_name in workbook.sheetnames:
                    if 'TEST ARTICLE LOG' in sheet_name.upper() and 'TEST RESULTS' in sheet_name.upper():
                        target_sheet = workbook[sheet_name]
                        print(f"🔍 DEBUG - Found target sheet: {sheet_name}")
                        break
                
                if target_sheet is None:
                    # Fallback: look for sheets containing 'TEST ARTICLE' or 'LOG'
                    for sheet_name in workbook.sheetnames:
                        if 'TEST ARTICLE' in sheet_name.upper() or 'LOG' in sheet_name.upper():
                            target_sheet = workbook[sheet_name]
                            print(f"🔍 DEBUG - Found fallback sheet: {sheet_name}")
                            break
                
                if target_sheet:
                    # Extract data from the worksheet
                    print(f"🔍 DEBUG - Extracting data from sheet: {target_sheet.title}")
                    print(f"🔍 DEBUG - Sheet dimensions: {target_sheet.max_row} rows x {target_sheet.max_column} cols")
                    
                    formatted_data += f"\n=== WORKSHEET: {target_sheet.title} ===\n"
                    
                    # Process all data from the sheet
                    max_rows = target_sheet.max_row  # Process all rows
                    max_cols = min(target_sheet.max_column, 20)  # Reasonable column limit
                    
                    print(f"🔍 DEBUG - Processing {max_rows} rows from worksheet")
                    
                    data_rows_found = 0
                    for row in range(1, max_rows + 1):
                        row_data = []
                        for col in range(1, max_cols + 1):
                            cell_value = target_sheet.cell(row=row, column=col).value
                            if cell_value is not None:
                                row_data.append(str(cell_value))
                            else:
                                row_data.append("")
                        
                        # Only include rows with some data
                        if any(cell.strip() for cell in row_data):
                            formatted_data += " | ".join(row_data) + "\n"
                            data_rows_found += 1
                    
                    print(f"🔍 DEBUG - Extracted {data_rows_found} data rows from sheet (total rows in file: {max_rows})")
                else:
                    print(f"🔍 DEBUG - No suitable worksheet found in file")
                
                # Cleanup temporary file
                os.unlink(temp_file_path)
                
            except Exception as e:
                print(f"🔍 DEBUG - Error processing file: {str(e)}")
                formatted_data += f"\n=== ERROR PROCESSING FILE: {str(e)} ===\n"
                continue
        
        if not formatted_data.strip():
            print("🔍 DEBUG - No formatted data found, returning default message")
            formatted_data = "No test article data found in uploaded files."
        
        print(f"🔍 DEBUG - Final formatted data length: {len(formatted_data)}")
        print(f"🔍 DEBUG - Final formatted data preview: {formatted_data[:300]}...")
        
        return formatted_data
    
    def _count_test_articles_in_content(self, excel_data: str) -> int:
        """
        Count the number of unique DUT based on unique Serial Numbers
        """
        lines = excel_data.split('\n')
        unique_serial_numbers = set()
        
        # Look for data lines with Serial Number patterns
        for line in lines:
            if '|' in line and line.strip() and not line.startswith('==='):
                # Skip header rows by checking for header keywords
                upper_line = line.upper()
                if any(header in upper_line for header in ['DUT PART NUMBER', 'PART NUMBER', 'SERIAL NUMBER', 'LOT NUMBER']):
                    continue
                
                # Split line and look for serial number (second column typically)
                parts = [p.strip() for p in line.split('|')]
                if len(parts) >= 2:
                    # Serial number is typically in second column
                    potential_serial = parts[1].strip()
                    
                    # Check if it looks like a serial number (numeric, reasonable length)
                    if potential_serial and potential_serial.isdigit() and len(potential_serial) >= 8:
                        unique_serial_numbers.add(potential_serial)
        
        dut_count = len(unique_serial_numbers)
        print(f"🔍 DEBUG - Found {dut_count} unique DUT Serial Numbers: {list(unique_serial_numbers)[:5]}{'...' if len(unique_serial_numbers) > 5 else ''}")
        return max(1, dut_count)  # Minimum 1
    
    async def _create_test_article_attachment(self, data_files: List[Any], report_config: dict) -> dict:
        """
        Create Raw Data attachment for reports with >10 test articles
        Uses the complete uploaded Excel file as attachment
        """
        import os
        
        try:
            # Find the Excel file
            for data_file in data_files:
                if not data_file or not hasattr(data_file, 'filename'):
                    continue
                    
                if data_file.filename.endswith(('.xlsx', '.xls')):
                    # Create attachment filename with new naming convention
                    report_num = report_config.get('report_number', 'RPT-XXX')
                    revision = report_config.get('revision', '001')
                    attachment_name = f"{report_num} rev{revision} Attachment - Raw Data.xlsx"
                    
                    # Save file content to attachments directory
                    attachments_dir = os.path.join(os.getcwd(), "attachments")
                    os.makedirs(attachments_dir, exist_ok=True)
                    
                    attachment_path = os.path.join(attachments_dir, attachment_name)
                    
                    # Read original Excel file content
                    await data_file.seek(0)
                    file_content = await data_file.read()
                    
                    # Save the complete Excel file as attachment
                    with open(attachment_path, 'wb') as f:
                        f.write(file_content)
                    
                    print(f"✅ Created raw data attachment: {attachment_name}")
                    
                    return {
                        "filename": attachment_name,
                        "path": attachment_path,
                        "size": os.path.getsize(attachment_path),
                        "original_filename": data_file.filename
                    }
                    
        except Exception as e:
            print(f"⚠️ Failed to create attachment: {str(e)}")
            
        return None
    
    def _create_fallback_section(self, device_config: Dict[str, Any]) -> str:
        """
        Create fallback section when AI processing fails
        """
        sterilized_text = "sterilized" if device_config.get("units_sterilized") else "not sterilized"
        
        if device_config.get("units_modified"):
            modification_text = f"Modifications were made: {device_config.get('modification_description', 'Details not specified')}"
        else:
            modification_text = "No modifications were made outside normal manufacturing processes."
        
        return f"""DEVICE UNDER TEST CONFIGURATION

The test units were {sterilized_text}. {modification_text}

[Test article details will be populated from uploaded Excel data]

| Part Number | Serial Number | Lot Number |
|-------------|---------------|------------|
| TBD | TBD | TBD |"""


class EquipmentUsedAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.6: Create Equipment Used Section
    Integrates with Excel Parser to process Equipment Log, Software Log, and Material Log data
    Uses AI to intelligently organize and present consumables and equipment information
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.agent_name = "EquipmentUsedAgent"
        self.target_sheets = ["EQUIPMENT LOG", "SOFTWARE LOG", "MATERIAL LOG"]
    
    async def process(
        self, 
        test_data: List[Dict[str, Any]], 
        report_config: Dict[str, Any] = None,
        calibration_verified: bool = True
    ) -> TaskResult:
        """
        Process equipment data from Excel Parser and generate intelligent Equipment Used section
        
        Args:
            test_data: List of test data dictionaries from Excel Parser
            report_config: Report configuration (number, revision, etc.)
            calibration_verified: Whether equipment calibration was verified
            
        Returns:
            TaskResult with AI-organized Equipment Used section content
        """
        try:
            print(f"🔧 {self.agent_name}: Processing equipment data with AI organization...")
            
            # Extract and organize data from Excel Parser
            parsed_data = self._extract_excel_parser_data(test_data)
            
            if not parsed_data or all(len(data) == 0 for data in parsed_data.values()):
                return TaskResult(
                    success=True,
                    content="No equipment, software, or material log data was found in the uploaded files.",
                    metadata={"agent_name": self.agent_name}
                )
            
            # Count items and generate summary
            equipment_count = len(parsed_data.get("EQUIPMENT LOG", []))
            software_count = len(parsed_data.get("SOFTWARE LOG", []))
            material_count = len(parsed_data.get("MATERIAL LOG", []))
            
            print(f"📊 Found: {equipment_count} equipment, {software_count} software, {material_count} material items")
            
            # Use AI to intelligently organize and format the data
            ai_organized_content = await self._ai_organize_equipment_data(
                parsed_data, calibration_verified, report_config
            )
            
            # Handle attachments if needed (>5 items in any category)
            needs_attachment = (equipment_count > 5 or software_count > 5 or material_count > 5)
            attachment_info = None
            
            if needs_attachment:
                attachment_info = await self._create_equipment_attachment(test_data, report_config)
                # Add attachment reference to AI content
                if attachment_info:
                    ai_organized_content += f"\n\nDetailed equipment, software, and material information is provided in {attachment_info.get('filename', 'Attachment B')}."
            
            print(f"✅ {self.agent_name}: Successfully generated AI-organized equipment section")
            
            result = TaskResult(
                success=True,
                content=ai_organized_content,
                metadata={
                    "agent_name": self.agent_name,
                    "equipment_count": equipment_count,
                    "software_count": software_count,
                    "material_count": material_count
                }
            )
            
            if attachment_info:
                result.attachments = [attachment_info]
            
            return result
            
        except Exception as e:
            error_msg = f"EquipmentUsedAgent processing failed: {str(e)}"
            print(f"❌ {error_msg}")
            return TaskResult(
                success=False,
                content="Unable to process equipment data at this time.",
                metadata={"agent_name": self.agent_name}, 
                error=error_msg
            )
    
    def _extract_excel_parser_data(self, test_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """Extract equipment data from Excel Parser results"""
        equipment_data = {sheet_name: [] for sheet_name in self.target_sheets}
        
        print(f"🔍 EquipmentUsedAgent: Extracting Excel Parser data for: {self.target_sheets}")
        
        for data_item in test_data:
            # Look for equipment data from Excel Parser
            if "equipment_data" in data_item and isinstance(data_item["equipment_data"], dict):
                # This is equipment data from Excel Parser
                for sheet_name, sheet_info in data_item["equipment_data"].items():
                    if sheet_name in self.target_sheets and isinstance(sheet_info, dict):
                        # Extract the actual data rows
                        if "data" in sheet_info and isinstance(sheet_info["data"], list):
                            for row in sheet_info["data"]:
                                if isinstance(row, dict) and any(str(v).strip() for v in row.values() if v is not None):
                                    equipment_data[sheet_name].append(row)
            
            # Also check for direct sheet data (fallback)
            elif "sheet_name" in data_item and data_item["sheet_name"] in self.target_sheets:
                if "content" in data_item and isinstance(data_item["content"], list):
                    for row in data_item["content"]:
                        if isinstance(row, dict) and any(str(v).strip() for v in row.values() if v is not None):
                            equipment_data[data_item["sheet_name"]].append(row)
        
        # Print summary results
        for sheet_name in self.target_sheets:
            print(f"📊 Found {len(equipment_data[sheet_name])} items in '{sheet_name}'")
            if equipment_data[sheet_name]:
                # Show sample data structure
                sample = equipment_data[sheet_name][0]
                print(f"   Sample columns: {list(sample.keys())[:5]}...")
        
        return equipment_data
    
    async def _ai_organize_equipment_data(
        self, 
        parsed_data: Dict[str, List[Dict[str, Any]]], 
        calibration_verified: bool,
        report_config: Dict[str, Any] = None
    ) -> str:
        """Use AI to intelligently organize and format equipment data"""
        
        # Prepare data summary for AI
        data_summary = self._prepare_data_summary(parsed_data)
        
        # Build AI prompt
        prompt = f"""You are an AI assistant helping to organize equipment, software, and material information for a DVT (Design Verification Test) report.

Equipment Data Summary:
{data_summary}

Calibration Status: {"All equipment was verified as calibrated" if calibration_verified else "Equipment calibration was not verified"}

Please organize this information into a professional Equipment Used section with the following STRICT formatting requirements:

1. Start with the calibration verification statement

2. Create THREE separate sections IN ORDER with clear headers:
   - "• Equipment Used" (for EQUIPMENT LOG data)
   - "• Software Used" (for SOFTWARE LOG data) 
   - "• Materials Used" (for MATERIAL LOG data)

3. For EACH section, follow this EXACT structure:
   a) Write the section header (e.g., "• Equipment Used")
   b) Write the status statement on the next line:
      - If no data: "No equipment/software/materials were recorded in the logs."
      - If data exists: "The equipment/software/materials log recorded the following items:"
   c) If data exists and ≤5 items: Place the table IMMEDIATELY after the status statement
   d) If data exists and >5 items: Add "Additional items are listed in the attachment."

4. TABLE PLACEMENT RULES (CRITICAL):
   - Equipment tables must appear ONLY under "• Equipment Used"
   - Software tables must appear ONLY under "• Software Used"
   - Materials tables must appear ONLY under "• Materials Used"
   - Do NOT place tables in the wrong sections
   - Each table must include ALL columns from the original data

5. CRITICAL REQUIREMENTS:
   - You MUST create all three sections regardless of data availability
   - Tables must be placed immediately under their corresponding section
   - Never mix table data between sections
   - Use actual data values, no placeholder text like "TBD"

Format everything in clean Markdown. Follow the section order and table placement rules exactly."""
        
        try:
            print(f"🤖 {self.agent_name}: Using AI to organize equipment data...")
            
            response = await self.client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {"role": "system", "content": "You are a technical writing assistant specializing in DVT reports and equipment documentation."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3
            )
            
            ai_content = response.choices[0].message.content.strip()
            print(f"✅ {self.agent_name}: AI successfully organized equipment data")
            
            return ai_content
            
        except Exception as e:
            print(f"⚠️ {self.agent_name}: AI organization failed, using fallback: {str(e)}")
            # Fallback to simple organization
            return self._fallback_organize_data(parsed_data, calibration_verified)
    
    def _prepare_data_summary(self, parsed_data: Dict[str, List[Dict[str, Any]]]) -> str:
        """Prepare a detailed summary of the equipment data for AI processing"""
        summary_parts = []
        
        for sheet_name, data_list in parsed_data.items():
            if not data_list:
                summary_parts.append(f"\n{sheet_name}: No data found")
                continue
                
            summary_parts.append(f"\n{sheet_name} ({len(data_list)} items):")
            
            # Show column structure
            if data_list:
                columns = list(data_list[0].keys())
                summary_parts.append(f"  Columns: {', '.join(columns)}")
                
                # Show ALL items with full details (not truncated)
                for i, item in enumerate(data_list):
                    item_details = []
                    for k, v in item.items():
                        if v and str(v).strip():
                            item_details.append(f"{k}: {str(v)}")
                    if item_details:
                        summary_parts.append(f"  Item {i+1}: {', '.join(item_details)}")
        
        return '\n'.join(summary_parts)
    
    def _fallback_organize_data(
        self, 
        parsed_data: Dict[str, List[Dict[str, Any]]], 
        calibration_verified: bool
    ) -> str:
        """Fallback method to organize data without AI"""
        content_parts = []
        
        # Add calibration statement
        if calibration_verified:
            content_parts.append("All equipment used in this testing was verified as calibrated at the time of use.")
        else:
            content_parts.append("Equipment calibration status was not verified at the time of use.")
        
        content_parts.append("")
        
        # Process each category explicitly
        categories = [
            ("EQUIPMENT LOG", "• Equipment Used", "equipment"),
            ("SOFTWARE LOG", "• Software Used", "software"), 
            ("MATERIAL LOG", "• Materials Used", "materials")
        ]
        
        for sheet_name, section_title, item_type in categories:
            content_parts.append(f"{section_title}")
            
            data_list = parsed_data.get(sheet_name, [])
            
            if not data_list:
                content_parts.append(f"No {item_type} were recorded in the {item_type} log.")
            else:
                content_parts.append(f"The {item_type} log recorded the following items:")
                
                if len(data_list) <= 5:
                    # Create inline table
                    table_md = self._create_markdown_table(sheet_name, data_list)
                    if table_md:
                        # Remove the section header from table since we added it above
                        table_lines = table_md.split('\n')
                        if table_lines and table_lines[0].startswith('###'):
                            table_lines = table_lines[1:]
                        content_parts.extend([""] + table_lines)
                else:
                    # Reference to attachment
                    content_parts.append(f"\nDetailed {item_type} information is provided in the attachment.")
            
            content_parts.append("")  # Add spacing between sections
        
        return '\n'.join(content_parts)

    def _generate_calibration_statement(self, calibration_verified: bool) -> str:
        """Generate calibration verification statement"""
        if calibration_verified:
            return "All equipment used in this testing was verified as calibrated at the time of use."
        else:
            return "Equipment calibration status was not verified at the time of use."
    
    def _create_markdown_table(self, sheet_name: str, data: List[Dict[str, Any]]) -> str:
        """Create markdown table from sheet data"""
        if not data:
            return ""
        
        # Get column headers from first row
        headers = list(data[0].keys())
        
        # Create table header
        table_lines = []
        
        # Add section header
        if sheet_name == "EQUIPMENT LOG":
            table_lines.append("• Equipment Used")
        elif sheet_name == "SOFTWARE LOG":
            table_lines.append("• Software Used")
        elif sheet_name == "MATERIAL LOG":
            table_lines.append("• Materials Used")
        
        # Add table header row
        header_row = "| " + " | ".join(headers) + " |"
        table_lines.append(header_row)
        
        # Add separator row
        separator_row = "|" + "|".join([" --- " for _ in headers]) + "|"
        table_lines.append(separator_row)
        
        # Add data rows
        for row in data:
            values = [str(row.get(header, "")).strip() for header in headers]
            data_row = "| " + " | ".join(values) + " |"
            table_lines.append(data_row)
        
        return "\n".join(table_lines)
    
    async def _create_equipment_attachment(
        self, 
        test_data: List[Dict[str, Any]], 
        report_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Create Attachment B for equipment logs using complete Excel file"""
        try:
            import os
            import shutil
            
            # Generate attachment filename
            report_num = report_config.get('report_number', 'RPT-XXX') if report_config else 'RPT-XXX'
            revision = report_config.get('revision', '001') if report_config else '001'
            attachment_name = f"{report_num} rev{revision} Attachment - Raw Data.xlsx"
            
            # Find the first Excel file from the uploaded test data
            source_file = None
            for data_item in test_data:
                file_path = data_item.get("file_path", "")
                if file_path and file_path.lower().endswith('.xlsx'):
                    source_file = file_path
                    break
            
            if not source_file or not os.path.exists(source_file):
                print(f"⚠️ No Excel source file found for equipment attachment")
                return None
            
            # Create attachments directory
            attachments_dir = "attachments"
            os.makedirs(attachments_dir, exist_ok=True)
            attachment_path = os.path.join(attachments_dir, attachment_name)
            
            # Copy the complete Excel file as attachment
            shutil.copy2(source_file, attachment_path)
            
            print(f"✅ Created equipment attachment: {attachment_name}")
            
            return {
                "filename": attachment_name,
                "path": attachment_path,
                "size": os.path.getsize(attachment_path),
                "type": "excel"
            }
                
        except Exception as e:
            print(f"⚠️ Failed to create equipment attachment: {str(e)}")
            return None

class TestMethodLossAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.11: Create Test Method Loss Investigations Section
    Specializes in analyzing TEST METHOD LOSSES worksheet and generating investigation reports
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("test_method_loss")
    
    async def create_test_method_loss_investigations(self, excel_data_dict: Dict[str, Any], 
                                                   report_config: Dict[str, str] = None) -> TaskResult:
        """
        Process Test Method Loss Investigations task
        
        Args:
            excel_data_dict: Dictionary containing Excel worksheet data
            report_config: Report configuration with document number and revision
            
        Returns:
            TaskResult with generated content or error message
        """
        try:
            print(f"\n🔍 Starting Task 4.11: Test Method Loss Investigations")
            
            if not report_config:
                report_config = {
                    'document_number': 'RPT-XXXXXX', 
                    'revision': 'XXX'
                }
            
            # Extract TEST METHOD LOSSES data
            test_method_loss_data = self._extract_test_method_loss_data(excel_data_dict)
            
            if not test_method_loss_data:
                print("🔍 [TASK 4.11 DEBUG] No TEST METHOD LOSSES worksheet found or no data available")
                return TaskResult(
                    success=True,
                    content=self._generate_no_losses_content(),
                    metadata={"message": "No test method losses data found"},
                    error=None
                )
            
            print(f"🔍 [TASK 4.11 DEBUG] Found {test_method_loss_data.get('total_investigations', 0)} investigations")
            
            # Process the data and generate content using AI
            processed_content = await self._generate_ai_content(test_method_loss_data, report_config)
            
            if not processed_content:
                return TaskResult(
                    success=False,
                    content="",
                    metadata={"test_method_loss_data": test_method_loss_data},
                    error="Failed to generate AI content for test method loss investigations"
                )
            
            print(f"✅ Task 4.11 completed successfully")
            
            return TaskResult(
                success=True,
                content=processed_content,
                metadata={"test_method_loss_data": test_method_loss_data},
                error=None
            )
            
        except Exception as e:
            print(f"❌ Error in Task 4.11: {str(e)}")
            return TaskResult(
                success=False,
                content="",
                metadata={},
                error=f"Error processing test method loss investigations: {str(e)}"
            )
    
    def _extract_test_method_loss_data(self, excel_data_dict: Dict[str, Any]) -> Dict[str, Any]:
        """Extract data from TEST METHOD LOSSES worksheet"""
        try:
            # Look for TEST METHOD LOSSES worksheet
            test_method_loss_sheet = None
            sheet_data = None
            
            print(f"🔍 [TASK 4.11 DEBUG] Excel data dict keys: {list(excel_data_dict.keys())}")
            
            # Try different possible sheet names
            possible_sheet_names = [
                "TEST METHOD LOSSES",
                "Test Method Losses", 
                "test method losses",
                "TEST_METHOD_LOSSES",
                "TestMethodLosses"
            ]
            
            for filename, file_data in excel_data_dict.items():
                print(f"🔍 [TASK 4.11 DEBUG] Processing file: {filename}")
                print(f"🔍 [TASK 4.11 DEBUG] File data type: {type(file_data)}")
                
                if isinstance(file_data, dict) and 'sheets' in file_data:
                    available_sheets = list(file_data['sheets'].keys())
                    print(f"🔍 [TASK 4.11 DEBUG] Available sheets: {available_sheets}")
                    
                    for sheet_name, data in file_data['sheets'].items():
                        print(f"🔍 [TASK 4.11 DEBUG] Checking sheet: '{sheet_name}'")
                        if sheet_name.upper() == "TEST METHOD LOSSES":
                            test_method_loss_sheet = sheet_name
                            sheet_data = data
                            print(f"🔍 [TASK 4.11 DEBUG] Found exact match: {sheet_name}")
                            break
                    if sheet_data:
                        break
            
            if not sheet_data:
                print("🔍 [TASK 4.11 DEBUG] No TEST METHOD LOSSES worksheet found")
                return {}
            
            print(f"🔍 [TASK 4.11 DEBUG] Found worksheet: {test_method_loss_sheet}")
            print(f"🔍 [TASK 4.11 DEBUG] Sheet data type: {type(sheet_data)}")
            print(f"🔍 [TASK 4.11 DEBUG] Sheet data length: {len(sheet_data) if sheet_data else 0}")
            
            # Process the sheet data
            investigations = []
            headers = []
            
            for i, row in enumerate(sheet_data):
                if i == 0:  # Header row
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    print(f"🔍 [TASK 4.11 DEBUG] Headers: {headers}")
                    continue
                
                if not any(cell for cell in row):  # Skip empty rows
                    continue
                
                # Create investigation entry
                investigation = {}
                for j, cell in enumerate(row):
                    if j < len(headers) and headers[j]:
                        investigation[headers[j]] = str(cell).strip() if cell else ""
                
                # Only include rows that have test method loss type
                protocol_note_type = investigation.get('PROTOCOL NOTE TYPE', '').lower()
                print(f"🔍 [TASK 4.11 DEBUG] Row {i} protocol note type: '{protocol_note_type}'")
                if 'test method loss' in protocol_note_type:
                    investigations.append(investigation)
                    print(f"🔍 [TASK 4.11 DEBUG] Added investigation: {investigation}")
            
            result = {
                'worksheet_name': test_method_loss_sheet,
                'total_investigations': len(investigations),
                'investigations': investigations,
                'headers': headers
            }
            
            print(f"🔍 [TASK 4.11 DEBUG] Processed {len(investigations)} valid investigations")
            return result
            
        except Exception as e:
            print(f"❌ Error extracting test method loss data: {str(e)}")
            return {}
    
    async def _generate_ai_content(self, test_method_loss_data: Dict[str, Any], report_config: Dict[str, str]) -> str:
        """Generate AI content for test method loss investigations"""
        try:
            if not test_method_loss_data.get('investigations'):
                return self._generate_no_losses_content()
            
            # Format the data for AI prompt
            formatted_data = self._format_data_for_ai(test_method_loss_data)
            print(f"🔍 [TASK 4.11 DEBUG] AI prompt data length: {len(formatted_data)}")
            
            # Generate AI prompt
            prompt = DVTPrompts.test_method_loss_prompt(formatted_data, report_config)
            print(f"🔍 [TASK 4.11 DEBUG] Generated AI prompt with {len(prompt)} characters")
            
            # Get AI response
            response = await self.generate_content(prompt, self.default_temperature)
            
            if response:
                formatted_content = self._format_final_content(response)
                print(f"🔍 [TASK 4.11 DEBUG] Generated content length: {len(formatted_content)}")
                return formatted_content
            else:
                print("⚠️ [TASK 4.11 DEBUG] No AI response received")
                return self._generate_no_losses_content()
                
        except Exception as e:
            print(f"❌ Error generating AI content: {str(e)}")
            return self._generate_no_losses_content()
    
    def _format_data_for_ai(self, test_method_loss_data: Dict[str, Any]) -> str:
        """Format extracted data for AI prompt"""
        investigations = test_method_loss_data.get('investigations', [])
        
        if not investigations:
            return "No test method loss investigations found."
        
        formatted_lines = ["TEST METHOD LOSSES Data:"]
        formatted_lines.append("=" * 50)
        
        for i, investigation in enumerate(investigations, 1):
            formatted_lines.append(f"\nInvestigation #{i}:")
            formatted_lines.append(f"- #: {investigation.get('#', '')}")
            formatted_lines.append(f"- PROTOCOL NOTE TYPE: {investigation.get('PROTOCOL NOTE TYPE', '')}")
            formatted_lines.append(f"- TEST UNIT SN or ID: {investigation.get('TEST UNIT SN or ID', '')}")
            formatted_lines.append(f"- PROTOCOL STEP / SECTION: {investigation.get('PROTOCOL STEP / SECTION', '')}")
            formatted_lines.append(f"- REPLACEMENT(S): {investigation.get('REPLACEMENT(S)', '')}")
            formatted_lines.append(f"- OBSERVATIONS: {investigation.get('OBSERVATIONS', '')}")
        
        return "\n".join(formatted_lines)
    
    def _generate_no_losses_content(self) -> str:
        """Generate content when no test method losses are found"""
        return """No test method loss occurred in the execution of this test."""
    
    def _format_final_content(self, ai_response: str) -> str:
        """Format the final content with proper structure"""
        if not ai_response.strip():
            return self._generate_no_losses_content()
        
        # Clean up the content
        content = ai_response.strip()
        
        # Remove any existing section headers that AI might have added
        if content.startswith("## TEST METHOD LOSS INVESTIGATIONS"):
            content = content.replace("## TEST METHOD LOSS INVESTIGATIONS", "").strip()
        if content.startswith("# TEST METHOD LOSS INVESTIGATIONS"):
            content = content.replace("# TEST METHOD LOSS INVESTIGATIONS", "").strip()
        
        return content


class DVTAgentOrchestrator:
    """
    Orchestrator for managing all DVT AI agents and task execution
    Handles the proper execution order and data flow between agents
    """
    
    def __init__(self, client, model_name: str = None):
        self.client = client
        self.model_name = model_name or AgentConfig.DEFAULT_MODEL
        
        # Initialize all specialized agents
        self.scope_agent = ScopeAndPurposeAgent(client, model_name)
        self.reference_agent = ReferenceAgent(client, model_name)
        self.procedure_agent = TestProcedureSummaryAgent(client, model_name)
        self.acronyms_agent = AcronymsDefinitionsAgent(client, model_name)
        self.device_config_agent = DeviceUnderTestAgent(client, model_name)
        self.equipment_agent = EquipmentUsedAgent(client, model_name)
        self.test_result_summary_agent = TestResultSummaryAgent(client, model_name)
    
    async def execute_tasks_4_1_to_4_4(self, protocol_content: str, protocol_number: str, 
                                      project_name: str, report_content: str = "") -> Dict[str, TaskResult]:
        """
        Execute AI tasks 4.1-4.4 in proper order
        
        Args:
            protocol_content: Protocol document content
            protocol_number: Protocol document number and revision
            project_name: Project name for the report
            report_content: Complete report content (for task 4.3, executed last)
            
        Returns:
            Dictionary with results for each task
        """
        
        results = {}
        
        # Task 4.1: Create Scope and Purpose
        print("🚀 Executing Task 4.1: Create Scope and Purpose")
        results["task_4_1"] = await self.scope_agent.create_scope_and_purpose(
            protocol_content, protocol_number, project_name
        )
        
        # Task 4.2: Create Reference Section  
        print("🚀 Executing Task 4.2: Create Reference Section")
        # Use protocol content to scan for document references
        scan_content = protocol_content + "\n" + report_content if report_content else protocol_content
        results["task_4_2"] = await self.reference_agent.create_reference_section(scan_content)
        
        # Task 4.4: Create Test Procedure Summary
        print("🚀 Executing Task 4.4: Create Test Procedure Summary")  
        results["task_4_4"] = await self.procedure_agent.create_test_procedure_summary(
            protocol_content, protocol_number
        )
        
        # Task 4.3: Create Acronyms & Definitions (EXECUTED LAST)
        print("🚀 Executing Task 4.3: Create Acronyms & Definitions (Final Step)")
        if report_content:
            results["task_4_3"] = await self.acronyms_agent.create_acronyms_and_definitions(report_content)
        else:
            # If no complete report yet, create placeholder
            results["task_4_3"] = TaskResult(
                success=True,
                content="Task 4.3 will be executed after complete report is generated",
                metadata={"status": "pending_complete_report"}
            )
        
        return results
    
    async def execute_final_acronyms_task(self, complete_report_content: str) -> TaskResult:
        """
        Execute Task 4.3 after complete report is generated
        
        Args:
            complete_report_content: Complete report content to scan for acronyms
            
        Returns:
            TaskResult with acronyms and definitions section
        """
        print("🚀 Executing Final Task 4.3: Create Acronyms & Definitions")
        return await self.acronyms_agent.create_acronyms_and_definitions(complete_report_content)
    
    async def execute_task_4_5(self, data_files: List[Any], device_config: Dict[str, Any], 
                              report_config: Dict[str, Any]) -> TaskResult:
        """
        Execute Task 4.5: Create Device Under Test Configuration Section
        
        Args:
            data_files: List of uploaded Excel files containing test article data
            device_config: User configuration (sterilized, modified, etc.)
            report_config: Report configuration (number, revision, etc.)
            
        Returns:
            TaskResult with device configuration section
        """
        print("🚀 Executing Task 4.5: Create Device Under Test Configuration Section")
        return await self.device_config_agent.create_device_configuration_section(
            data_files, device_config, report_config
        )
    
    async def execute_task_4_7(self, doc_data: Dict[str, Any], excel_data: Dict[str, Any], 
                              test_data_type: str, protocol_content: str = "", 
                              ai_friendly_format: Optional[str] = None) -> TaskResult:
        """
        Execute Task 4.7: Create Test Result Summary
        
        Args:
            doc_data: Parsed document data from doc parser
            excel_data: Parsed Excel data 
            test_data_type: "attribute" or "variable"
            protocol_content: Protocol content as backup
            ai_friendly_format: AI-friendly formatted document text
            
        Returns:
            TaskResult with test result summary table
        """
        print("🚀 Executing Task 4.7: Create Test Result Summary")
        return await self.test_result_summary_agent.create_test_result_summary(
            doc_data, excel_data, test_data_type, protocol_content, ai_friendly_format
        )
    
    async def test_ai_connection(self) -> str:
        """Test AI connection with all agents"""
        if not self.client:
            return ErrorConfig.ERROR_MESSAGES["ai_not_configured"]
        
        try:
            test_result = await self.scope_agent.generate_content(
                DVTPrompts.ai_connection_test_prompt(), 
                temperature=AgentConfig.get_temperature("connection_test")
            )
            return f"✅ {test_result}"
        except Exception as e:
            return f"❌ AI test failed: {str(e)}"


class TestResultSummaryAgent(BaseDVTAgent):
    """
    AI Agent for Task 4.7: Create Test Result Summary
    Specializes in generating test result summary tables based on data type
    """
    
    def __init__(self, client, model_name: str = None):
        super().__init__(client, model_name)
        self.default_temperature = AgentConfig.get_temperature("test_result_summary")
    
    async def create_test_result_summary(self, 
                                       doc_data: Dict[str, Any],
                                       excel_data: Dict[str, Any], 
                                       test_data_type: str,
                                       protocol_content: Optional[str] = None,
                                       ai_friendly_format: Optional[str] = None) -> TaskResult:
        """
        Generate test result summary table based on data type
        
        Args:
            doc_data: Parsed document data
            excel_data: Parsed Excel data
            test_data_type: "attribute" or "variable"
            protocol_content: Backup protocol content if needed
            
        Returns:
            TaskResult with generated table
        """
        try:
            # Debug: Check data types
            print(f"🔍 [TASK 4.7 DEBUG] doc_data type: {type(doc_data)}")
            print(f"🔍 [TASK 4.7 DEBUG] excel_data type: {type(excel_data)}")
            print(f"🔍 [TASK 4.7 DEBUG] protocol_content type: {type(protocol_content)}")
            print(f"🔍 [TASK 4.7 DEBUG] ai_friendly_format type: {type(ai_friendly_format)}")
            
            if isinstance(doc_data, dict):
                print(f"🔍 [TASK 4.7 DEBUG] doc_data keys: {list(doc_data.keys())}")
            elif isinstance(doc_data, list):
                print(f"🔍 [TASK 4.7 DEBUG] doc_data length: {len(doc_data)}")
            
            if isinstance(excel_data, dict):
                print(f"🔍 [TASK 4.7 DEBUG] excel_data keys: {list(excel_data.keys())}")
            elif isinstance(excel_data, list):
                print(f"🔍 [TASK 4.7 DEBUG] excel_data length: {len(excel_data)}")
            
            # Step 1: Get Acceptance Criteria data
            acceptance_criteria_data = self._get_acceptance_criteria(doc_data, protocol_content, ai_friendly_format)
            if not acceptance_criteria_data:
                print("⚠️ No Acceptance Criteria found, generating default template")
                # Generate default template instead of failing
                return self._generate_default_table(test_data_type, excel_data)
            
            # Step 2: Extract key information using AI
            criteria_info = await self._extract_criteria_info(acceptance_criteria_data)
            if not criteria_info:
                print("⚠️ Failed to extract criteria info, generating default template")
                # Generate default template instead of failing
                return self._generate_default_table(test_data_type, excel_data)
            
            # Step 3: Get Excel data (TEST ARTICLE LOG & DEFECTIVE UNITS)
            excel_stats = self._get_excel_statistics(excel_data)
            
            # Step 4: Generate table based on data type
            table_content = self._generate_table(criteria_info, excel_stats, test_data_type)
            
            return TaskResult(
                success=True,
                content=table_content,
                metadata={
                    "test_data_type": test_data_type,
                    "criteria_count": len(criteria_info),
                    "excel_stats": excel_stats
                }
            )
            
        except Exception as e:
            return TaskResult(
                success=False,
                content="",
                metadata={},
                error=f"Error in create_test_result_summary: {str(e)}"
            )
    
    def _get_acceptance_criteria(self, doc_data: Dict[str, Any], protocol_content: Optional[str] = None, ai_friendly_format: Optional[str] = None) -> str:
        """Let AI extract Acceptance Criteria from the full document data using AI-friendly format"""
        print(f"🔍 Preparing document for AI analysis...")
        
        # Use AI-friendly format if provided, otherwise fallback to JSON
        if ai_friendly_format:
            doc_text = ai_friendly_format
            print(f"📄 Using provided AI-friendly format")
        else:
            # Fallback to JSON format
            doc_text = json.dumps(doc_data, indent=2, ensure_ascii=False)
            print(f"📄 Using JSON format as fallback")
        
        # If protocol_content is available, include it as well
        if protocol_content:
            doc_text += "\n\nProtocol Content:\n" + protocol_content
        
        print(f"📄 Document data prepared for AI analysis ({len(doc_text)} characters)")
        return doc_text
    
    async def _extract_criteria_info(self, acceptance_criteria_data: str) -> List[Dict[str, str]]:
        """Use AI to extract REQ ID, Acceptance Criteria description, and Confidence/Reliability"""
        try:
            prompt = DVTPrompts.test_result_summary_extraction_prompt(acceptance_criteria_data)
            response = await self.generate_content(prompt)
            
            # Parse AI response to extract structured data
            criteria_list = self._parse_criteria_response(response)
            return criteria_list
            
        except Exception as e:
            print(f"❌ Error extracting criteria info: {str(e)}")
            return []
    
    def _parse_criteria_response(self, response: str) -> List[Dict[str, str]]:
        """Parse AI response into structured criteria data"""
        criteria_list = []
        
        try:
            # Look for JSON structure in response
            import re
            json_match = re.search(r'\[.*\]', response, re.DOTALL)
            if json_match:
                criteria_data = json.loads(json_match.group())
                for item in criteria_data:
                    if all(key in item for key in ["req_id", "acceptance_criteria", "confidence_reliability"]):
                        criteria_list.append(item)
        except:
            # Fallback: parse text format
            lines = response.split('\n')
            current_criteria = {}
            
            for line in lines:
                line = line.strip()
                if line.startswith("REQ ID:"):
                    if current_criteria:
                        criteria_list.append(current_criteria)
                    current_criteria = {"req_id": line.replace("REQ ID:", "").strip()}
                elif line.startswith("Acceptance Criteria:"):
                    current_criteria["acceptance_criteria"] = line.replace("Acceptance Criteria:", "").strip()
                elif line.startswith("Confidence/Reliability:"):
                    current_criteria["confidence_reliability"] = line.replace("Confidence/Reliability:", "").strip()
            
            if current_criteria:
                criteria_list.append(current_criteria)
        
        return criteria_list
    
    def _get_excel_statistics(self, excel_data: Dict[str, Any]) -> Dict[str, int]:
        """Extract statistics from Excel data - use parsed statistics if available"""
        stats = {
            "initial_sample_size": 0,
            "test_method_losses": 0,
            "actual_sample_size": 0,
            "defective_units": 0
        }
        
        debug_info = []
        debug_info.append(f"🔍 [TASK 4.7 DEBUG] Excel data structure: {list(excel_data.keys()) if isinstance(excel_data, dict) else type(excel_data)}")
        
        # Handle nested excel data structure
        actual_data = excel_data
        if isinstance(excel_data, dict) and len(excel_data) == 1:
            # If excel_data is {'filename.xlsx': actual_data}, unwrap it
            first_key = list(excel_data.keys())[0]
            if isinstance(excel_data[first_key], dict):
                actual_data = excel_data[first_key]
                debug_info.append(f"📁 Unwrapped data from file: {first_key}")
        
        # Check if actual_data is dict before calling .items()
        if isinstance(actual_data, dict):
            for key, value in actual_data.items():
                if isinstance(value, (list, dict)):
                    debug_info.append(f"   - {key}: {type(value)} with {len(value)} items")
                else:
                    debug_info.append(f"   - {key}: {value}")
            
            # Also check if we have the 'sheets' key and what sheets are available
            if 'sheets' in actual_data:
                sheets_data = actual_data['sheets']
                if isinstance(sheets_data, dict):
                    debug_info.append(f"📄 Available sheet names: {list(sheets_data.keys())}")
                    # Show sample data for each sheet
                    for sheet_name, sheet_data in sheets_data.items():
                        if isinstance(sheet_data, list) and len(sheet_data) > 0:
                            debug_info.append(f"   📋 {sheet_name}: {len(sheet_data)} rows")
                        
        else:
            debug_info.append(f"   - actual_data is {type(actual_data)} with {len(actual_data)} items")

        # Write debug info to file
        try:
            with open("task_4_7_debug.log", "w", encoding="utf-8") as f:
                f.write("\n".join(debug_info))
                f.write(f"\n\nFull actual_data content:\n{json.dumps(actual_data, indent=2, default=str, ensure_ascii=False)}")
        except Exception as e:
            print(f"Warning: Could not write debug file: {e}")
        
        # Also print to console
        for line in debug_info:
            print(line)

        try:
            # Only process if actual_data is a dictionary
            if not isinstance(actual_data, dict):
                print(f"⚠️ actual_data is {type(actual_data)}, expected dict. Cannot extract statistics.")
                return stats
            
            # Check if we need to look inside 'sheets' key
            search_data = actual_data
            if 'sheets' in actual_data and isinstance(actual_data['sheets'], dict):
                search_data = actual_data['sheets']
                print(f"🔍 Searching in sheets data with keys: {list(search_data.keys())}")
                
            # Try multiple possible key names for TEST ARTICLE LOG
            test_article_data = None
            for key_name in ["TEST ARTICLE LOG & TEST RESULTS", "TEST ARTICLE LOG", "Test Article Log", "test_data"]:
                if key_name in search_data:
                    test_article_data = search_data[key_name]
                    print(f"✅ Found test article data under key: '{key_name}'")
                    break
                    
            if test_article_data and isinstance(test_article_data, list):
                stats["initial_sample_size"] = len(test_article_data) - 1  # Subtract header row
                print(f"📊 Initial sample size: {stats['initial_sample_size']}")
                
                # Count TML and FAIL from actual test results
                test_method_losses_count = 0
                defective_units_count = 0
                
                print(f"🔍 Analyzing {len(test_article_data)} rows for TML and FAIL results")
                for i, row in enumerate(test_article_data[1:], 1):  # Skip header row
                    if isinstance(row, (list, tuple)) and len(row) > 5:
                        test_result_column = str(row[5]) if len(row) > 5 else ""
                        test_result_upper = test_result_column.upper().strip()
                        
                        if "TML" in test_result_upper:
                            test_method_losses_count += 1
                            print(f"   ✅ Found TML in row {i}: {test_result_column}")
                        elif "FAIL" in test_result_upper:
                            defective_units_count += 1
                            print(f"   ✅ Found FAIL in row {i}: {test_result_column}")
                
                stats["test_method_losses"] = test_method_losses_count
                stats["defective_units"] = defective_units_count
                print(f"📊 Test method losses: {stats['test_method_losses']}")
                print(f"📊 Defective units: {stats['defective_units']}")
                
            # Calculate actual sample size
            stats["actual_sample_size"] = stats["initial_sample_size"] - stats["test_method_losses"]
            print(f"📊 Actual sample size: {stats['actual_sample_size']}")
            
        except Exception as e:
            print(f"⚠️ Warning extracting Excel statistics: {str(e)}")
        
        print(f"📊 Final statistics: {stats}")
        return stats
    
    def _generate_default_table(self, test_data_type: str, excel_data: Dict[str, Any]) -> TaskResult:
        """Generate default table when Acceptance Criteria is not found"""
        excel_stats = self._get_excel_statistics(excel_data)
        
        # Create default criteria info
        default_criteria = [{
            "req_id": "TBD",
            "acceptance_criteria": "TBD (Please update with actual acceptance criteria)",
            "confidence_reliability": "90%/90%"
        }]
        
        table_content = self._generate_table(default_criteria, excel_stats, test_data_type)
        
        return TaskResult(
            success=True,
            content=table_content + "\n\n*Note: Default template generated. Please update with actual Acceptance Criteria data.*",
            metadata={
                "test_data_type": test_data_type,
                "criteria_count": 1,
                "excel_stats": excel_stats,
                "is_default": True
            }
        )
    
    def _generate_table(self, criteria_info: List[Dict[str, str]], excel_stats: Dict[str, int], test_data_type: str) -> str:
        """Generate Word table based on data type"""
        if test_data_type.lower() == "attribute":
            return self._generate_attribute_table(criteria_info, excel_stats)
        else:
            return self._generate_variable_table(criteria_info, excel_stats)
    
    def _generate_attribute_table(self, criteria_info: List[Dict[str, str]], excel_stats: Dict[str, int]) -> str:
        """Generate table for attribute data type"""
        from scipy.stats import binom
        
        table_rows = []
        
        for criteria in criteria_info:
            # Calculate Actual Confidence/Reliability
            try:
                confidence_rel = criteria.get("confidence_reliability", "90%/90%")
                parts = confidence_rel.split("/")
                confidence_level = float(parts[0].replace("%", "")) / 100  # e.g., 0.90 for 90%
                reliability_target = float(parts[1].replace("%", "")) / 100  # e.g., 0.90 for 90%
                
                num_failures = excel_stats["defective_units"]
                sample_size = excel_stats["actual_sample_size"]
                
                if sample_size > 0 and num_failures >= 0:
                    # Calculate actual confidence based on observed data
                    actual_confidence = 1 - (num_failures / sample_size)
                    
                    # Format: Actual Confidence / Target Reliability
                    actual_conf_rel = f"{actual_confidence:.2%} / {reliability_target:.2%}"
                else:
                    actual_conf_rel = "TBD"
            except Exception as e:
                print(f"Warning: Error calculating confidence: {e}")
                actual_conf_rel = "TBD"
            
            row = [
                criteria.get("req_id", ""),
                criteria.get("acceptance_criteria", ""),
                criteria.get("confidence_reliability", ""),
                str(excel_stats["initial_sample_size"]),
                str(excel_stats["test_method_losses"]),
                str(excel_stats["actual_sample_size"]),
                str(excel_stats["defective_units"]),
                actual_conf_rel
            ]
            table_rows.append(row)
        
        # Generate Word table format
        headers = [
            "REQ ID", "Acceptance Criteria", "Confidence / Reliability",
            "Initial Sample Size", "Test Method Losses", "Actual Sample Size",
            "Defective Units", "Actual Confidence/ Reliability"
        ]
        
        return self._format_word_table(headers, table_rows)
    
    def _generate_variable_table(self, criteria_info: List[Dict[str, str]], excel_stats: Dict[str, int]) -> str:
        """Generate table for variable data type"""
        table_rows = []
        
        for criteria in criteria_info:
            row = [
                criteria.get("req_id", ""),
                criteria.get("acceptance_criteria", ""),
                str(excel_stats["initial_sample_size"]),
                str(excel_stats["test_method_losses"]),
                str(excel_stats["actual_sample_size"]),
                str(excel_stats["defective_units"]),
                "",  # Tolerance Interval - leave empty as requested
                ""   # Confidence / Reliability - leave empty as requested
            ]
            table_rows.append(row)
        
        # Generate Word table format
        headers = [
            "REQ ID", "Acceptance Criteria", "Initial Sample Size",
            "Test Method Losses", "Actual Sample Size", "Defective Units",
            "Tolerance Interval", "Confidence / Reliability"
        ]
        
        return self._format_word_table(headers, table_rows)
    
    def _format_word_table(self, headers: List[str], rows: List[List[str]]) -> str:
        """Format table in Word document style using markdown table format"""
        # Create markdown table format
        table_content = "\n"
        
        # Add header row
        header_row = "| " + " | ".join(headers) + " |\n"
        separator_row = "|" + "|".join([" --- " for _ in headers]) + "|\n"
        
        table_content += header_row
        table_content += separator_row
        
        # Add data rows
        for row in rows:
            data_row = "| " + " | ".join(row) + " |\n"
            table_content += data_row
        
        table_content += "\n"
        return table_content


class ProtocolDeviationsAgent(BaseDVTAgent):
    """
    AI Agent for processing protocol deviations data and generating formatted deviation reports
    Converts Excel deviation data into standardized deviation documentation
    """
    
    def __init__(self, client, model_name: Optional[str] = None):
        super().__init__(client, model_name)
        self.task_name = "protocol_deviations"
        self.default_temperature = AgentConfig.get_temperature(self.task_name)
    
    async def process_deviations(self, deviations_text: str) -> TaskResult:
        """
        Process deviations data and generate formatted deviation report
        
        Args:
            deviations_text: Formatted deviation data from Excel parser
            
        Returns:
            TaskResult with formatted deviation content
        """
        if not deviations_text or deviations_text.strip() == "No deviations data found.":
            return TaskResult(
                success=True,
                content="No deviations.",
                metadata={"deviation_count": 0, "processing_method": "fallback"}
            )
        
        try:
            # Get prompt from DVTPrompts
            prompt = DVTPrompts.get_protocol_deviations_prompt(deviations_text)
            
            # Generate content using AI
            raw_content = await self.generate_content(
                prompt, 
                temperature=self.default_temperature
            )
            
            if not raw_content:
                return TaskResult(
                    success=False,
                    content="No deviations.",
                    metadata={"error": "AI generation returned empty content"},
                    error="Failed to generate deviation content"
                )
            
            # Validate and post-process the generated content
            processed_content = self._validate_and_format(raw_content)
            
            # Count deviations for metadata
            deviation_count = self._count_deviations(processed_content)
            
            return TaskResult(
                success=True,
                content=processed_content,
                metadata={
                    "deviation_count": deviation_count,
                    "processing_method": "ai_generated",
                    "validation_passed": True
                }
            )
            
        except Exception as e:
            print(f"❌ Error processing deviations: {str(e)}")
            return TaskResult(
                success=False,
                content="No deviations.",
                metadata={"error": str(e)},
                error=f"Deviation processing failed: {str(e)}"
            )
    
    def _validate_and_format(self, raw_content: str) -> str:
        """
        Validate and format the AI-generated deviation content
        
        Args:
            raw_content: Raw content from AI generation
            
        Returns:
            Validated and formatted deviation content
        """
        content = raw_content.strip()
        
        # If no deviations found, return standard message
        if not content or "no deviations" in content.lower():
            return "No deviations."
        
        # Validate deviation numbering format
        content = self._fix_deviation_numbering(content)
        
        # Clean up formatting
        content = self._clean_formatting(content)
        
        # Validate required elements are present
        if not self._has_required_elements(content):
            print("⚠️ Warning: Generated deviation content may be missing required elements")
        
        return content
    
    def _fix_deviation_numbering(self, content: str) -> str:
        """Fix and standardize deviation numbering"""
        lines = content.split('\n')
        processed_lines = []
        deviation_counter = 1
        
        for line in lines:
            line = line.strip()
            
            # Look for deviation headers and standardize them
            if self._is_deviation_header(line):
                # Replace with standardized format
                standardized_header = f"DEVIATION #{deviation_counter}: {self._extract_deviation_title(line)}"
                processed_lines.append(standardized_header)
                deviation_counter += 1
            else:
                processed_lines.append(line)
        
        return '\n'.join(processed_lines)
    
    def _is_deviation_header(self, line: str) -> bool:
        """Check if line is a deviation header"""
        line_upper = line.upper()
        patterns = [
            r'DEVIATION\s*#?\d+', 
            r'DEVIATION\s+\d+',
            r'DEV\s*#?\d+',
            r'^\d+\.',  # Simple numbering like "1."
        ]
        
        for pattern in patterns:
            if re.search(pattern, line_upper):
                return True
        return False
    
    def _extract_deviation_title(self, line: str) -> str:
        """Extract deviation title from header line"""
        # Remove deviation numbering and extract title
        line = re.sub(r'DEVIATION\s*#?\d*:?', '', line, flags=re.IGNORECASE).strip()
        line = re.sub(r'DEV\s*#?\d*:?', '', line, flags=re.IGNORECASE).strip()
        line = re.sub(r'^\d+\.?\s*', '', line).strip()
        
        # If title is empty or too short, provide a generic title
        if len(line) < 5:
            return "Protocol deviation"
        
        return line
    
    def _clean_formatting(self, content: str) -> str:
        """Clean up formatting issues"""
        # Remove extra blank lines
        content = re.sub(r'\n\s*\n\s*\n', '\n\n', content)
        
        # Ensure proper spacing after colons
        content = re.sub(r':\s*([^\s])', r': \1', content)
        
        # Clean up any markdown artifacts that shouldn't be in Word
        content = re.sub(r'\*\*(.*?)\*\*', r'\1', content)  # Remove bold markdown
        content = re.sub(r'\*(.*?)\*', r'\1', content)      # Remove italic markdown
        
        return content.strip()
    
    def _has_required_elements(self, content: str) -> bool:
        """Check if content has required deviation elements"""
        content_lower = content.lower()
        
        # Check for key elements
        has_deviation_number = bool(re.search(r'deviation\s*#\d+', content_lower))
        has_impact_discussion = any(keyword in content_lower for keyword in [
            'impact', 'affect', 'result', 'consequence', 'influence'
        ])
        has_description = len(content) > 50  # Basic length check
        
        return has_deviation_number and has_impact_discussion and has_description
    
    def _count_deviations(self, content: str) -> int:
        """Count the number of deviations in the content"""
        if "no deviations" in content.lower():
            return 0
        
        # Count DEVIATION # patterns
        pattern = r'DEVIATION\s*#\d+'
        matches = re.findall(pattern, content, flags=re.IGNORECASE)
        return len(matches)


class DefectiveUnitInvestigationsAgent(BaseDVTAgent):
    """
    AI Agent for processing defective unit investigation data and generating formatted investigation reports
    Converts Excel defective unit data into standardized investigation documentation
    """
    
    def __init__(self, client, model_name: Optional[str] = None):
        super().__init__(client, model_name)
        self.task_name = "defective_unit_investigations"
        self.default_temperature = AgentConfig.get_temperature("protocol_deviations")  # Use same temp as deviations
    
    async def process_defective_units(self, defective_units_text: str) -> TaskResult:
        """
        Process defective units data and generate formatted investigation report
        
        Args:
            defective_units_text: Formatted defective unit data from Excel parser
            
        Returns:
            TaskResult with formatted defective unit investigations content
        """
        if not defective_units_text or defective_units_text.strip() == "No defective units data found.":
            return TaskResult(
                success=True,
                content="No defective unit in the execution of this test.",
                metadata={"investigation_count": 0, "processing_method": "fallback"}
            )
        
        try:
            # Get prompt from DVTPrompts
            prompt = DVTPrompts.get_defective_unit_investigations_prompt(defective_units_text)
            
            # Generate content using AI
            raw_content = await self.generate_content(
                prompt, 
                temperature=self.default_temperature
            )
            
            if not raw_content:
                return TaskResult(
                    success=False,
                    content="No defective unit in the execution of this test.",
                    metadata={"error": "AI generation returned empty content"},
                    error="Failed to generate defective unit content"
                )
            
            # Validate and post-process the generated content
            processed_content = self._validate_and_format(raw_content)
            
            # Count investigations for metadata
            investigation_count = self._count_investigations(processed_content)
            
            return TaskResult(
                success=True,
                content=processed_content,
                metadata={
                    "investigation_count": investigation_count,
                    "processing_method": "ai_generated",
                    "validation_passed": True
                }
            )
            
        except Exception as e:
            print(f"❌ Error processing defective units: {str(e)}")
            return TaskResult(
                success=False,
                content="No defective unit in the execution of this test.",
                metadata={"error": str(e)},
                error=f"Defective unit processing failed: {str(e)}"
            )
    
    def _validate_and_format(self, raw_content: str) -> str:
        """
        Validate and format the AI-generated defective unit content
        
        Args:
            raw_content: Raw content from AI generation
            
        Returns:
            Validated and formatted defective unit content
        """
        content = raw_content.strip()
        
        # If no defective units found, return standard message
        if not content or "no defective unit" in content.lower():
            return "No defective unit in the execution of this test."
        
        # Validate investigation numbering format
        content = self._fix_investigation_numbering(content)
        
        # Clean up formatting
        content = self._clean_formatting(content)
        
        # Validate required elements are present
        if not self._has_required_elements(content):
            print("⚠️ Warning: Generated defective unit content may be missing required elements")
        
        return content
    
    def _fix_investigation_numbering(self, content: str) -> str:
        """Fix and standardize investigation numbering"""
        lines = content.split('\n')
        processed_lines = []
        investigation_counter = 1
        
        for line in lines:
            line = line.strip()
            
            # Look for investigation headers and standardize them
            if self._is_investigation_header(line):
                # Replace with standardized format
                standardized_header = f"DEFECTIVE UNIT INVESTIGATION #{investigation_counter}"
                processed_lines.append(standardized_header)
                investigation_counter += 1
            else:
                processed_lines.append(line)
        
        return '\n'.join(processed_lines)
    
    def _is_investigation_header(self, line: str) -> bool:
        """Check if line is an investigation header"""
        line_upper = line.upper()
        patterns = [
            r'DEFECTIVE\s+UNIT\s+INVESTIGATION\s*#?\d*',
            r'INVESTIGATION\s*#?\d+',
            r'DEFECTIVE\s+UNIT\s*#?\d+',
            r'^\d+\.',  # Simple numbering like "1."
        ]
        
        for pattern in patterns:
            if re.search(pattern, line_upper):
                return True
        return False
    
    def _clean_formatting(self, content: str) -> str:
        """Clean up formatting issues"""
        # Remove extra blank lines
        content = re.sub(r'\n\s*\n\s*\n', '\n\n', content)
        
        # Ensure proper spacing after colons
        content = re.sub(r':\s*([^\s])', r': \1', content)
        
        # Clean up any markdown artifacts that shouldn't be in Word
        content = re.sub(r'\*\*(.*?)\*\*', r'\1', content)  # Remove bold markdown
        content = re.sub(r'\*(.*?)\*', r'\1', content)      # Remove italic markdown
        
        return content.strip()
    
    def _has_required_elements(self, content: str) -> bool:
        """Check if content has required investigation elements"""
        if "no defective unit" in content.lower():
            return True
            
        content_lower = content.lower()
        
        # Check for key elements
        has_investigation_number = bool(re.search(r'defective\s+unit\s+investigation\s*#\d+', content_lower))
        has_specification_info = any(keyword in content_lower for keyword in [
            'specification', 'requirement', 'failed', 'communicate', 'not met'
        ])
        has_unit_info = any(keyword in content_lower for keyword in [
            'serial', 'unit', 'article', 'test article'
        ])
        has_investigation_info = any(keyword in content_lower for keyword in [
            'review', 'investigation', 'root cause', 'manufacturing', 'test method'
        ])
        
        return has_investigation_number and has_specification_info and has_unit_info and has_investigation_info
    
    def _count_investigations(self, content: str) -> int:
        """Count the number of investigations in the content"""
        if "no defective unit" in content.lower():
            return 0
        
        # Count DEFECTIVE UNIT INVESTIGATION # patterns
        pattern = r'DEFECTIVE\s+UNIT\s+INVESTIGATION\s*#\d+'
        matches = re.findall(pattern, content, flags=re.IGNORECASE)
        return len(matches)


class ConclusionAgent(BaseDVTAgent):
    """
    AI Agent for generating conclusion section based on test results and scope
    Analyzes test results against acceptance criteria and provides summary conclusions
    """
    
    def __init__(self, client, model_name: str = "gemini-2.0-flash-exp"):
        super().__init__(client, model_name)
        self.task_type = "conclusion"
        self.default_temperature = AgentConfig.get_temperature("conclusion")
    
    async def generate_conclusion(self, test_results_content: str, scope_content: str) -> TaskResult:
        """
        Generate conclusion section from test results and scope content
        
        Args:
            test_results_content: Content from [BK_TEST_RESULT_SUMMARY] section
            scope_content: Content from [BK_SCOPE_TEXT] section
        
        Returns:
            TaskResult containing the generated conclusion
        """
        if not test_results_content or not scope_content:
            return TaskResult(
                success=False,
                content="",
                error="Missing test results or scope content for conclusion generation",
                metadata={"missing_inputs": True}
            )
        
        try:
            # Get specialized prompt for conclusion generation
            from .ai_prompts import DVTPrompts
            prompts = DVTPrompts()
            
            prompt = prompts.get_conclusion_prompt(test_results_content, scope_content)
            
            # Generate conclusion with AI using configured temperature
            content = await self.generate_content(prompt, temperature=self.default_temperature)
            
            if content:
                # Validate the generated conclusion
                validation_result = self._validate_conclusion(content)
                
                return TaskResult(
                    success=True,
                    content=content,
                    error="",
                    metadata={
                        "input_length": len(test_results_content) + len(scope_content),
                        "output_length": len(content),
                        "validation_passed": validation_result["valid"],
                        "validation_notes": validation_result.get("notes", [])
                    }
                )
            else:
                return TaskResult(
                    success=False,
                    content="",
                    error="AI failed to generate conclusion content",
                    metadata={"generation_failed": True}
                )
                
        except Exception as e:
            return TaskResult(
                success=False,
                content="",
                error=f"Error in conclusion generation: {str(e)}",
                metadata={"exception": str(e)}
            )
    
    def _validate_conclusion(self, content: str) -> Dict[str, Any]:
        """
        Validate the generated conclusion content
        
        Args:
            content: Generated conclusion text
        
        Returns:
            Dictionary with validation results
        """
        validation_result = {
            "valid": True,
            "notes": []
        }
        
        # Check for key elements
        content_lower = content.lower()
        
        # Check for acceptance criteria reference
        if "acceptance criteria" not in content_lower and "requirement" not in content_lower:
            validation_result["notes"].append("Missing acceptance criteria reference")
        
        # Check for protocol reference
        if "protocol" not in content_lower and "pt-" not in content_lower:
            validation_result["notes"].append("Missing protocol reference")
        
        # Check for test results format (units/total)
        import re
        if not re.search(r'\d+\s*(out\s*of|/)\s*\d+', content):
            validation_result["notes"].append("Missing test results format (x out of y units)")
        
        # Check for confidence level analysis (new requirement)
        confidence_keywords = ["confidence level", "actual confidence", "required confidence", "confidence"]
        if not any(keyword in content_lower for keyword in confidence_keywords):
            validation_result["notes"].append("Missing confidence level analysis")
        
        # Check for data type analysis (attribute/variable)
        data_type_keywords = ["attribute data", "variable data", "tolerance interval", "specification limit"]
        if any(keyword in content_lower for keyword in data_type_keywords):
            validation_result["notes"].append("Good: Includes data type analysis")
        
        # Check for explicit acceptance criteria evaluation
        criteria_evaluation = ["criteria was met", "criteria is met", "meets the requirements", "exceeds the required"]
        if any(phrase in content_lower for phrase in criteria_evaluation):
            validation_result["notes"].append("Good: Explicit criteria evaluation")
        
        # Check minimum length
        if len(content.strip()) < 150:  # Increased from 100 due to more detailed requirements
            validation_result["valid"] = False
            validation_result["notes"].append("Conclusion too short for detailed analysis")
        
        return validation_result
